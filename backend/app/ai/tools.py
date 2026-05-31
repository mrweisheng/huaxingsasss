"""
Agent 工具执行器
调用现有 Service 层实现业务操作
"""
import base64
import json
import os
import shutil
from uuid import uuid4
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

from datetime import timedelta

import httpx
import structlog
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.config import settings
from app.core.chinese import search_variants
from app.models.customer import Customer
from app.models.contract import Contract
from app.models.payment import Payment
from app.models.user import User
from app.ai.prompts import RECEIPT_ANALYSIS_PROMPT, CONTRACT_ANALYSIS_PROMPT, GENERAL_ANALYSIS_PROMPT
from app.services.contract_service import ContractService
from app.services.customer_service import CustomerService
from app.services.payment_service import PaymentService
from app.utils.file_utils import calculate_file_hash

logger = structlog.get_logger()


def _escape_ilike(keyword: str) -> str:
    """转义 ILIKE 通配符，防止 SQL 注入"""
    return keyword.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


class ToolExecutor:
    """Agent 工具执行器，每个方法返回 JSON 字符串"""

    def __init__(self, db: Session, user: User):
        self.db = db
        self.user = user

    def _can_access_contract(self, contract: Contract) -> bool:
        if self.user.role == "admin":
            return True
        if self.user.role == "expense":
            return True  # expense 可查看所有合同（用于关联支出）
        return contract.sales_person_id == self.user.id

    def _is_admin(self) -> bool:
        return self.user.role == "admin"

    def _can_view_income(self) -> bool:
        return self.user.role in ("admin", "income")

    def _can_view_expense(self) -> bool:
        return self.user.role in ("admin", "expense")

    def _can_create_contract(self) -> bool:
        return self.user.role in ("admin", "income")

    def _contract_to_dict(self, c: Contract) -> dict:
        return {
            "id": c.id,
            "contract_number": c.contract_number,
            "title": c.title,
            "business_type": c.business_type,
            "business_description": c.business_description,
            "customer_name": c.customer.name if c.customer else None,
            "currency": c.currency,
            "total_amount": float(c.total_amount) if c.total_amount else 0,
            "paid_amount": float(c.paid_amount) if c.paid_amount else 0,
            "remaining_amount": float(c.remaining_amount) if c.remaining_amount else 0,
            "total_amount_in_cny": float(c.total_amount_in_cny) if c.total_amount_in_cny else None,
            "paid_amount_in_cny": float(c.paid_amount_in_cny) if c.paid_amount_in_cny else 0,
            "total_expense": float(c.total_expense) if c.total_expense else 0,
            "total_expense_in_cny": float(c.total_expense_in_cny) if c.total_expense_in_cny else 0,
            "status": c.status,
            "wechat_group": c.wechat_group,
            "signed_date": str(c.signed_date) if c.signed_date else None,
            "end_date": str(c.end_date) if c.end_date else None,
            "payment_stats": {
                "total": getattr(c, 'payment_total_count', 0),
                "paid": getattr(c, 'paid_count', 0),
                "expense_count": getattr(c, 'expense_count', 0),
            },
        }

    def _payment_to_dict(self, p: Payment) -> dict:
        return {
            "id": p.id,
            "contract_id": p.contract_id,
            "installment_number": p.installment_number,
            "installment_name": p.installment_name,
            "type": p.type,
            "payee_name": p.payee_name,
            "currency": p.currency,
            "amount": float(p.amount) if p.amount else 0,
            "paid_amount": float(p.paid_amount) if p.paid_amount else 0,
            "exchange_rate": float(p.exchange_rate) if p.exchange_rate else None,
            "amount_in_cny": float(p.amount_in_cny) if p.amount_in_cny else None,
            "paid_amount_in_cny": float(p.paid_amount_in_cny) if p.paid_amount_in_cny else None,
            "due_date": str(p.due_date) if p.due_date else None,
            "paid_date": str(p.paid_date) if p.paid_date else None,
            "payment_method": p.payment_method,
            "status": p.status,
            "notes": p.notes,
            "receipt_image_path": p.receipt_image_path,
            "receipt_data": p.receipt_data,
        }

    # ── 查询工具 ──

    def search_customers(
        self,
        name: Optional[str] = None,
        phone: Optional[str] = None,
        wechat_group: Optional[str] = None,
        limit: int = 10,
    ) -> str:
        if not (name or phone or wechat_group):
            return json.dumps({"error": "请至少提供一个搜索条件（name/phone/wechat_group）"}, ensure_ascii=False)

        query = self.db.query(Customer).filter(Customer.is_deleted == False)

        if name:
            variants = search_variants(name)
            escaped = [_escape_ilike(v) for v in variants]
            query = query.filter(or_(*[Customer.name.ilike(f"%{v}%") for v in escaped]))
        if phone:
            query = query.filter(Customer.phone.ilike(f"%{_escape_ilike(phone)}%"))
        if wechat_group:
            variants = search_variants(wechat_group)
            escaped = [_escape_ilike(v) for v in variants]
            query = query.filter(or_(*[Customer.wechat_group_name.ilike(f"%{v}%") for v in escaped]))

        customers = query.limit(limit).all()
        results = []
        for c in customers:
            contract_count = self.db.query(Contract).filter(
                Contract.customer_id == c.id,
                Contract.is_deleted == False,
            ).count()
            results.append({
                "id": c.id,
                "name": c.name,
                "contact_person": c.contact_person,
                "phone": c.phone,
                "wechat_group_name": c.wechat_group_name,
                "contract_count": contract_count,
            })

        return json.dumps({"customers": results, "total": len(results)}, ensure_ascii=False)

    def search_contracts(self, **kwargs) -> str:
        sales_person_id = None
        if self.user.role == "income":
            sales_person_id = self.user.id

        try:
            items, total = ContractService.get_contracts(
                self.db,
                page=kwargs.get("page", 1),
                per_page=kwargs.get("per_page", 10),
                status=kwargs.get("status"),
                customer_id=kwargs.get("customer_id"),
                customer_name=kwargs.get("customer_name"),
                keyword=kwargs.get("keyword"),
                date_from=kwargs.get("date_from"),
                date_to=kwargs.get("date_to"),
                sales_person_id=sales_person_id,
                contract_number=kwargs.get("contract_number"),
            )
        except Exception as e:
            return json.dumps({"error": str(e)}, ensure_ascii=False)

        contracts = [self._contract_to_dict(c) for c in items]
        return json.dumps({"contracts": contracts, "total": total}, ensure_ascii=False)

    def get_contract_detail(self, contract_id: int) -> str:
        contract = ContractService.get_contract(self.db, contract_id)
        if not contract:
            return json.dumps({"error": f"合同不存在：{contract_id}"}, ensure_ascii=False)
        if not self._can_access_contract(contract):
            return json.dumps({"error": "无权查看该合同"}, ensure_ascii=False)

        result = self._contract_to_dict(contract)
        result["customer_id"] = contract.customer_id
        result["sales_person_id"] = contract.sales_person_id
        result["remarks"] = contract.remarks

        try:
            # 按角色过滤：income 只看收入，expense 只看支出，admin 看全部
            if self.user.role == "income":
                type_filter = "income"
            elif self.user.role == "expense":
                type_filter = "expense"
            else:
                type_filter = None

            payment_data = PaymentService.get_contract_payments(self.db, contract_id, type_filter=type_filter)

            if type_filter != "expense":
                result["income"] = payment_data["income"]
            if type_filter != "income":
                result["expense"] = payment_data["expense"]
            # 利润只有 admin 能看
            if self._is_admin():
                result["profit_in_cny"] = payment_data["profit_in_cny"]
        except Exception:
            result["income"] = {"payments": []}
            result["expense"] = {"payments": []}

        return json.dumps(result, ensure_ascii=False)

    def get_customer_contracts(self, customer_id: int) -> str:
        sales_person_id = None
        if self.user.role == "income":
            sales_person_id = self.user.id

        items, total = ContractService.get_contracts(
            self.db,
            customer_id=customer_id,
            sales_person_id=sales_person_id,
            per_page=50,
        )
        contracts = [self._contract_to_dict(c) for c in items]
        return json.dumps({"contracts": contracts, "total": total}, ensure_ascii=False)

    def query_payments(self, **kwargs) -> str:
        query = self.db.query(Payment).filter(Payment.is_deleted == False)

        if kwargs.get("contract_id"):
            query = query.filter(Payment.contract_id == kwargs["contract_id"])

        if kwargs.get("status"):
            query = query.filter(Payment.status == kwargs["status"])

        if kwargs.get("overdue_only"):
            query = query.filter(
                Payment.due_date < date.today(),
                Payment.status == "pending",
            )

        if kwargs.get("type"):
            query = query.filter(Payment.type == kwargs["type"])

        # 角色权限：income 只看收入+自己合同，expense 只看支出+自己创建的
        if self.user.role == "income":
            query = query.filter(Payment.type == "income")
            query = query.join(Contract).filter(Contract.sales_person_id == self.user.id)
        elif self.user.role == "expense":
            query = query.filter(Payment.type == "expense")
            query = query.filter(Payment.created_by == self.user.id)

        page = kwargs.get("page", 1)
        per_page = kwargs.get("per_page", 20)
        total = query.count()
        payments = query.order_by(Payment.created_at.desc()) \
            .offset((page - 1) * per_page).limit(per_page).all()

        results = [self._payment_to_dict(p) for p in payments]

        for r, p in zip(results, payments):
            if p.contract:
                r["contract_number"] = p.contract.contract_number
                if p.contract.customer:
                    r["customer_name"] = p.contract.customer.name

        return json.dumps({"payments": results, "total": total}, ensure_ascii=False)

    # ── 动作工具 ──

    def create_customer(self, **kwargs) -> str:
        """创建客户。如果同名+同电话/邮箱的客户已存在，返回已有客户。"""
        if not self._can_create_contract():
            return json.dumps({"error": "当前角色无权创建客户"}, ensure_ascii=False)

        name = kwargs.get("name", "").strip()
        if not name:
            return json.dumps({"error": "客户姓名不能为空"}, ensure_ascii=False)

        phone = kwargs.get("phone", "").strip() or None
        email = kwargs.get("email", "").strip() or None

        if not phone and not email:
            return json.dumps({"error": "电话和邮箱至少填写一项"}, ensure_ascii=False)

        try:
            customer, created = CustomerService.create_or_get(
                db=self.db,
                name=name,
                phone=phone,
                email=email,
                contact_person=kwargs.get("contact_person"),
                id_card_number=kwargs.get("id_card_number"),
                business_license=kwargs.get("business_license"),
                address=kwargs.get("address"),
                wechat_group_name=kwargs.get("wechat_group_name"),
                remarks=kwargs.get("remarks"),
                created_by=self.user.id,
            )
            logger.info("create_customer: name=%s, id=%d, created=%s", name, customer.id, created)
            return json.dumps({
                "success": True,
                "customer": {
                    "id": customer.id,
                    "name": customer.name,
                    "phone": customer.phone,
                    "email": customer.email,
                    "wechat_group_name": customer.wechat_group_name,
                },
                "created": created,
                "message": "客户创建成功" if created else f"客户已存在（ID: {customer.id}）",
            }, ensure_ascii=False)
        except Exception as e:
            logger.exception("create_customer failed")
            return json.dumps({"error": f"创建客户失败: {str(e)}"}, ensure_ascii=False)

    def update_customer(self, **kwargs) -> str:
        """更新已有客户信息（电话、证件号等）"""
        if not self._can_create_contract():
            return json.dumps({"error": "当前角色无权修改客户"}, ensure_ascii=False)

        customer_id = kwargs.get("customer_id")
        if not customer_id:
            return json.dumps({"error": "缺少 customer_id"}, ensure_ascii=False)

        customer = self.db.query(Customer).filter(
            Customer.id == customer_id, Customer.is_deleted == False
        ).first()
        if not customer:
            return json.dumps({"error": f"客户不存在: {customer_id}"}, ensure_ascii=False)

        updatable = ["phone", "email", "id_card_number", "wechat_group_name", "address", "remarks"]
        updated = {}
        for field in updatable:
            val = kwargs.get(field)
            if val is not None:
                target_field = "id_card_number_encrypted" if field == "id_card_number" else field
                setattr(customer, target_field, val)
                updated[field] = val

        if not updated:
            return json.dumps({"error": "没有需要更新的字段"}, ensure_ascii=False)

        try:
            self.db.commit()
            self.db.refresh(customer)
            return json.dumps({
                "success": True,
                "customer": {
                    "id": customer.id,
                    "name": customer.name,
                    "phone": customer.phone,
                    "email": customer.email,
                    "wechat_group_name": customer.wechat_group_name,
                },
                "message": f"客户信息已更新: {list(updated.keys())}",
            }, ensure_ascii=False)
        except Exception as e:
            logger.exception("update_customer failed")
            return json.dumps({"error": f"更新客户失败: {str(e)}"}, ensure_ascii=False)

    @staticmethod
    def _guess_extension(content: bytes) -> str:
        """通过文件头判断扩展名"""
        if content[:4] == b"%PDF":
            return ".pdf"
        if content[:3] == b"\xff\xd8\xff":
            return ".jpg"
        if content[:4] == b"\x89PNG":
            return ".png"
        if content[:4] == b"GIF8":
            return ".gif"
        if content[:4] == b"RIFF" and len(content) > 11 and content[8:12] == b"WEBP":
            return ".webp"
        return ".bin"

    def _auto_create_payments_from_terms(
        self,
        contract: Contract,
        contract_data_raw: dict,
        receipt_data: dict = None,
        receipt_file_path: str = None,
    ) -> list:
        logger.info(
            "自动付款创建开始: contract_id=%d, contract_data类型=%s",
            contract.id, type(contract_data_raw).__name__,
        )
        if not isinstance(contract_data_raw, dict):
            logger.warning("自动付款创建跳过: contract_data不是dict（%s）", type(contract_data_raw).__name__)
            return []

        payment_terms = contract_data_raw.get("payment_terms", [])
        logger.info(
            "payment_terms数量=%d, 内容=%s",
            len(payment_terms),
            json.dumps(payment_terms, ensure_ascii=False)[:500],
        )
        if not payment_terms:
            logger.info("自动付款创建跳过: 无payment_terms")
            return []

        paid_keywords = ["已付", "已缴纳", "付清", "已到账", "已收", "已支付"]

        receipt_amount = None
        if receipt_data and isinstance(receipt_data, dict):
            try:
                receipt_amount = float(receipt_data.get("amount", 0))
            except (TypeError, ValueError):
                pass

        receipt_matched = False
        auto_payments = []
        for idx, term in enumerate(payment_terms, 1):
            # 优先使用结构化 is_paid 字段，回退到关键词匹配（兼容旧数据）
            is_paid_field = term.get("is_paid")
            condition = (term.get("condition") or "").lower()
            is_paid_by_keywords = any(kw in condition for kw in paid_keywords)
            is_paid_term = (is_paid_field is True) or (is_paid_field is None and is_paid_by_keywords)

            logger.info(
                "条款[%d]: name=%s, amount=%s, is_paid字段=%s, condition=%s, 关键词匹配=%s → is_paid=%s",
                idx, term.get("name"), term.get("amount"), is_paid_field,
                condition[:50], is_paid_by_keywords, is_paid_term,
            )

            if not is_paid_term:
                continue

            try:
                term_amount = float(term.get("amount", 0))
            except (TypeError, ValueError):
                continue

            if term_amount <= 0:
                continue

            matched_receipt = None
            if (
                not receipt_matched
                and receipt_file_path
                and receipt_amount
                and abs(term_amount - receipt_amount) < 1
            ):
                matched_receipt = receipt_file_path
                receipt_matched = True

            try:
                installment_number = PaymentService.get_next_installment_number(
                    self.db, contract.id, "income"
                )
                payment = PaymentService.create_payment_with_exchange_rate(
                    db=self.db,
                    contract_id=contract.id,
                    installment_number=installment_number,
                    currency=contract.currency,
                    amount=Decimal(str(term_amount)),
                    paid_date=contract.signed_date or date.today(),
                    payment_method="unknown",
                    receipt_image_path=matched_receipt,
                    notes="合同标注已付，待补充凭证" if not matched_receipt else "合同标注已付，已关联凭证",
                    created_by=self.user.id,
                    type="income",
                )
                if term.get("name"):
                    payment.installment_name = term["name"]
                if matched_receipt and receipt_data:
                    payment.receipt_data = receipt_data
                if term.get("name") or (matched_receipt and receipt_data):
                    self.db.commit()
                    self.db.refresh(payment)

                auto_payments.append({
                    "payment_id": payment.id,
                    "installment_number": idx,
                    "installment_name": term.get("name"),
                    "amount": term_amount,
                    "currency": contract.currency,
                    "status": payment.status,
                })
                logger.info(
                    "自动创建付款: contract_id=%d, term=%s, amount=%s, receipt=%s → status=%s",
                    contract.id, term.get("name"), term_amount,
                    "有" if matched_receipt else "无",
                    payment.status,
                )
            except Exception as e:
                logger.warning("自动创建付款失败: term=%s, error=%s", term, e)
                auto_payments.append({
                    "error": str(e),
                    "installment_name": term.get("name"),
                    "amount": term.get("amount"),
                })

        return auto_payments

    def create_contract(self, **kwargs) -> str:
        """为客户创建合同记录。合同编号自动生成。"""
        if not self._can_create_contract():
            return json.dumps({"error": "当前角色无权创建合同"}, ensure_ascii=False)

        customer_id = kwargs.get("customer_id")
        contract_data_raw = kwargs.get("contract_data", {})
        logger.info(
            "create_contract: customer_id=%s, file_id=%s, contract_data类型=%s, keys=%s",
            customer_id, kwargs.get("file_id"),
            type(contract_data_raw).__name__,
            list(contract_data_raw.keys()) if isinstance(contract_data_raw, dict) else "N/A",
        )
        file_id = kwargs.get("file_id")

        if not customer_id:
            return json.dumps({"error": "缺少 customer_id，请先创建或查找客户"}, ensure_ascii=False)
        if not file_id:
            return json.dumps({"error": "缺少 file_id，无法关联原始文件"}, ensure_ascii=False)

        # 验证客户存在
        customer = self.db.query(Customer).filter(
            Customer.id == customer_id, Customer.is_deleted == False
        ).first()
        if not customer:
            return json.dumps({"error": f"客户不存在: {customer_id}"}, ensure_ascii=False)

        # 处理文件
        temp_file_path = os.path.join(settings.TEMP_UPLOAD_DIR, file_id)
        file_hash = None
        original_file_path = f"agent_upload/{file_id}"

        if os.path.exists(temp_file_path):
            with open(temp_file_path, "rb") as f:
                content = f.read()
            file_hash = calculate_file_hash(content)

            # 基于文件 hash 检测重复
            existing = self.db.query(Contract).filter(
                Contract.file_hash == file_hash,
                Contract.is_deleted == False,
            ).first()
            if existing:
                return json.dumps({
                    "success": False,
                    "error": f"该文件已创建过合同（编号: {existing.contract_number}, ID: {existing.id}）",
                    "existing_contract_id": existing.id,
                }, ensure_ascii=False)

            # 复制到正式合同目录
            contract_number = ContractService.generate_contract_number()
            year_month = datetime.now().strftime("%Y/%m")
            target_dir = Path(settings.CONTRACT_UPLOAD_DIR) / year_month
            target_dir.mkdir(parents=True, exist_ok=True)

            ext = self._guess_extension(content)
            target_filename = f"{contract_number}{ext}"
            target_path = target_dir / target_filename
            shutil.copy2(temp_file_path, str(target_path))

            original_file_path = str(Path(year_month) / target_filename)
        else:
            contract_number = ContractService.generate_contract_number()

        # 解析日期
        signed_date = None
        if kwargs.get("signed_date"):
            try:
                signed_date = date.fromisoformat(kwargs["signed_date"])
            except ValueError:
                pass

        # 构建 schema 并创建
        try:
            from app.schemas.contract import ContractCreate

            contract_create = ContractCreate(
                contract_number=contract_number,
                title=kwargs.get("title"),
                business_type=kwargs.get("business_type"),
                business_description=kwargs.get("business_description"),
                customer_id=customer_id,
                currency=kwargs.get("currency", "CNY"),
                total_amount=Decimal(str(kwargs.get("total_amount", 0))),
                original_file_path=original_file_path,
                file_hash=file_hash,
                signed_date=signed_date,
                wechat_group=kwargs.get("wechat_group"),
                status="active",
            )

            contract = ContractService.create_contract(
                db=self.db,
                contract_data=contract_create,
                sales_person_id=self.user.id,
            )

            # 写入 contract_data JSON
            contract.contract_data = {
                "source": "agent",
                "file_id": file_id,
                "business_type": kwargs.get("business_type"),
                "business_description": kwargs.get("business_description"),
                **(contract_data_raw if isinstance(contract_data_raw, dict) else {}),
            }
            # 存储合同全文（用于知识库问答）
            if isinstance(contract_data_raw, dict) and contract_data_raw.get("full_text"):
                contract.contract_text = contract_data_raw["full_text"]
            self.db.commit()
            self.db.refresh(contract)

            auto_payments = self._auto_create_payments_from_terms(
                contract, contract_data_raw, kwargs.get("receipt_data"), kwargs.get("receipt_file_path")
            )
            logger.info(
                "create_contract完成: contract_id=%d, auto_payments=%s",
                contract.id, json.dumps(auto_payments, ensure_ascii=False)[:500],
            )

            return json.dumps({
                "success": True,
                "contract": {
                    "id": contract.id,
                    "contract_number": contract.contract_number,
                    "customer_name": customer.name,
                    "customer_id": customer_id,
                    "title": contract.title,
                    "currency": contract.currency,
                    "total_amount": float(contract.total_amount),
                    "status": contract.status,
                    "wechat_group": contract.wechat_group,
                    "signed_date": str(contract.signed_date) if contract.signed_date else None,
                },
                "auto_payments": auto_payments,
            }, ensure_ascii=False)
        except Exception as e:
            logger.exception("create_contract failed")
            return json.dumps({"error": f"创建合同失败: {str(e)}"}, ensure_ascii=False)

    def update_contract(self, **kwargs) -> str:
        """更新合同信息（如微信群、备注等）"""
        if not self._can_create_contract():
            return json.dumps({"error": "当前角色无权修改合同"}, ensure_ascii=False)

        contract_id = kwargs.get("contract_id")
        if not contract_id:
            return json.dumps({"error": "缺少 contract_id"}, ensure_ascii=False)

        contract = ContractService.get_contract(self.db, contract_id)
        if not contract:
            return json.dumps({"error": f"合同不存在：{contract_id}"}, ensure_ascii=False)
        if not self._can_access_contract(contract):
            return json.dumps({"error": "无权修改该合同"}, ensure_ascii=False)

        updatable_fields = ["wechat_group", "remarks", "title", "business_description"]
        updates = {}
        for field in updatable_fields:
            if field in kwargs and kwargs[field] is not None:
                updates[field] = kwargs[field]

        if not updates:
            return json.dumps({"error": "没有需要更新的字段"}, ensure_ascii=False)

        try:
            from app.schemas.contract import ContractUpdate
            contract_update = ContractUpdate(**updates)
            updated = ContractService.update_contract(self.db, contract_id, contract_update)
            if not updated:
                return json.dumps({"error": "更新失败"}, ensure_ascii=False)

            return json.dumps({
                "success": True,
                "contract": self._contract_to_dict(updated),
            }, ensure_ascii=False)
        except Exception as e:
            logger.exception("update_contract failed")
            return json.dumps({"error": f"更新合同失败: {str(e)}"}, ensure_ascii=False)

    def create_payment(self, **kwargs) -> str:
        """创建付款记录（收入类型）。"""
        if not self._can_view_income():
            return json.dumps({"error": "当前角色无权创建收入记录"}, ensure_ascii=False)

        required = ["contract_id", "installment_number", "amount", "currency", "paid_date"]
        missing = [r for r in required if not kwargs.get(r)]
        if missing:
            return json.dumps({"error": f"缺少必填参数: {', '.join(missing)}"}, ensure_ascii=False)

        if self.user.role == "income":
            contract = self.db.query(Contract).filter(Contract.id == kwargs["contract_id"]).first()
            if not contract or contract.sales_person_id != self.user.id:
                return json.dumps({"error": "无权操作该合同的付款"}, ensure_ascii=False)

        try:
            receipt_path = kwargs.get("receipt_image_path")
            logger.info(
                "Agent创建付款: contract_id=%s, installment=%s, amount=%s %s, receipt=%s, notes=%s",
                kwargs["contract_id"], kwargs["installment_number"],
                kwargs["amount"], kwargs.get("currency"),
                receipt_path or "无",
                kwargs.get("notes", "无"),
            )
            payment = PaymentService.create_payment_with_exchange_rate(
                db=self.db,
                contract_id=kwargs["contract_id"],
                installment_number=kwargs["installment_number"],
                currency=kwargs["currency"],
                amount=Decimal(str(kwargs["amount"])),
                paid_date=date.fromisoformat(kwargs["paid_date"]),
                payment_method=kwargs.get("payment_method", "unknown"),
                receipt_image_path=kwargs.get("receipt_image_path"),
                notes=kwargs.get("notes"),
                created_by=self.user.id,
                type="income",
            )
            self.db.refresh(payment)
            if kwargs.get("installment_name"):
                payment.installment_name = kwargs["installment_name"]
            if kwargs.get("receipt_data"):
                payment.receipt_data = kwargs["receipt_data"]
            if kwargs.get("installment_name") or kwargs.get("receipt_data"):
                self.db.commit()
                self.db.refresh(payment)
            if payment.contract and payment.contract.customer:
                payment.contract.customer  # ensure loaded
            result = self._payment_to_dict(payment)
            result["contract_number"] = payment.contract.contract_number if payment.contract else None
            result["customer_name"] = payment.contract.customer.name if payment.contract and payment.contract.customer else None
            return json.dumps({"success": True, "payment": result}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"error": str(e)}, ensure_ascii=False)

    def create_expense(self, **kwargs) -> str:
        """为合同创建支出记录（向第三方付款）。"""
        if not self._can_view_expense():
            return json.dumps({"error": "当前角色无权创建支出记录"}, ensure_ascii=False)

        required = ["contract_id", "amount", "currency", "paid_date", "payee_name"]
        missing = [r for r in required if not kwargs.get(r)]
        if missing:
            return json.dumps({"error": f"缺少必填参数: {', '.join(missing)}"}, ensure_ascii=False)

        try:
            # 自动计算 installment_number
            installment_number = PaymentService.get_next_installment_number(
                self.db, kwargs["contract_id"], "expense"
            )

            payment = PaymentService.create_payment_with_exchange_rate(
                db=self.db,
                contract_id=kwargs["contract_id"],
                installment_number=installment_number,
                currency=kwargs["currency"],
                amount=Decimal(str(kwargs["amount"])),
                paid_date=date.fromisoformat(kwargs["paid_date"]),
                payment_method=kwargs.get("payment_method", "unknown"),
                receipt_image_path=kwargs.get("receipt_image_path"),
                notes=kwargs.get("notes"),
                created_by=self.user.id,
                type="expense",
                payee_name=kwargs["payee_name"],
            )
            self.db.refresh(payment)
            if kwargs.get("installment_name"):
                payment.installment_name = kwargs["installment_name"]
            if kwargs.get("receipt_data"):
                payment.receipt_data = kwargs["receipt_data"]
            if kwargs.get("installment_name") or kwargs.get("receipt_data"):
                self.db.commit()
                self.db.refresh(payment)
            result = self._payment_to_dict(payment)
            result["contract_number"] = payment.contract.contract_number if payment.contract else None
            return json.dumps({"success": True, "payment": result}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"error": str(e)}, ensure_ascii=False)

    def update_payment(self, **kwargs) -> str:
        """更新付款记录的备注、凭证、付款方式等信息"""
        payment_id = kwargs.get("payment_id")
        if not payment_id:
            return json.dumps({"error": "缺少 payment_id"}, ensure_ascii=False)

        payment = self.db.query(Payment).filter(Payment.id == payment_id).first()
        if not payment:
            return json.dumps({"error": f"付款记录不存在：{payment_id}"}, ensure_ascii=False)

        # 按 type 权限检查
        if payment.type == "expense":
            if not self._can_view_expense():
                return json.dumps({"error": "当前角色无权更新支出记录"}, ensure_ascii=False)
        else:
            if not self._can_view_income():
                return json.dumps({"error": "当前角色无权更新收入记录"}, ensure_ascii=False)

        # income 角色校验合同所有权
        if self.user.role == "income" and payment.contract and payment.contract.sales_person_id != self.user.id:
            return json.dumps({"error": "无权操作该合同的付款"}, ensure_ascii=False)

        # 可更新字段白名单
        updatable_fields = ["notes", "payment_method", "receipt_image_path", "receipt_data", "installment_name", "paid_date"]
        updates = {f: kwargs[f] for f in updatable_fields if kwargs.get(f) is not None}

        if not updates:
            return json.dumps({"error": "没有需要更新的字段"}, ensure_ascii=False)

        try:
            from app.schemas.payment import PaymentUpdate
            payment_update = PaymentUpdate(**updates)
            updated = PaymentService.update_payment(self.db, payment_id, payment_update)
            if not updated:
                return json.dumps({"error": "更新失败"}, ensure_ascii=False)

            result = self._payment_to_dict(updated)
            return json.dumps({"success": True, "payment": result}, ensure_ascii=False)
        except Exception as e:
            logger.exception("update_payment failed")
            return json.dumps({"error": f"更新付款失败: {str(e)}"}, ensure_ascii=False)

    def match_receipt(self, **kwargs) -> str:
        receipt_data = kwargs.get("receipt_data", {})
        if not isinstance(receipt_data, dict):
            return json.dumps({"error": "receipt_data 必须是JSON对象"}, ensure_ascii=False)

        payer_name = receipt_data.get("payer_name", "")
        receipt_amount = None
        try:
            receipt_amount = float(receipt_data.get("amount", 0))
        except (TypeError, ValueError):
            pass
        receipt_currency = receipt_data.get("currency", "")

        customer_name_hint = kwargs.get("customer_name", "").strip()

        candidates = []

        if payer_name or customer_name_hint:
            search_name = customer_name_hint or payer_name
            variants = search_variants(search_name)
            escaped = [_escape_ilike(v) for v in variants]
            customers = self.db.query(Customer).filter(
                Customer.is_deleted == False,
                or_(*[Customer.name.ilike(f"%{v}%") for v in escaped]),
            ).limit(5).all()

            for customer in customers:
                contracts = self.db.query(Contract).filter(
                    Contract.customer_id == customer.id,
                    Contract.is_deleted == False,
                ).all()
                for contract in contracts:
                    if not self._can_access_contract(contract):
                        continue
                    pending_payments = self.db.query(Payment).filter(
                        Payment.contract_id == contract.id,
                        Payment.type == "income",
                        Payment.status == "pending",
                        Payment.is_deleted == False,
                    ).order_by(Payment.installment_number).all()

                    for p in pending_payments:
                        score = 0
                        match_reasons = []
                        if customer_name_hint:
                            score += 30
                            match_reasons.append("客户名指定")
                        else:
                            score += 20
                            match_reasons.append("客户名匹配")
                        if receipt_amount and p.amount:
                            diff = abs(float(p.amount) - receipt_amount)
                            if diff < 1:
                                score += 50
                                match_reasons.append("金额完全匹配")
                            elif diff / max(float(p.amount), receipt_amount) < 0.05:
                                score += 30
                                match_reasons.append("金额近似匹配")
                        if receipt_currency and p.currency == receipt_currency:
                            score += 10
                            match_reasons.append("币种匹配")

                        candidates.append({
                            "payment_id": p.id,
                            "contract_id": contract.id,
                            "contract_number": contract.contract_number,
                            "customer_name": customer.name,
                            "business_type": contract.business_type,
                            "business_description": contract.business_description,
                            "installment_number": p.installment_number,
                            "installment_name": p.installment_name,
                            "amount": float(p.amount) if p.amount else 0,
                            "currency": p.currency,
                            "status": p.status,
                            "paid_date": str(p.paid_date) if p.paid_date else None,
                            "score": score,
                            "match_reason": "、".join(match_reasons),
                        })

        if not candidates and receipt_amount:
            pending_query = self.db.query(Payment).filter(
                Payment.status == "pending",
                Payment.type == "income",
                Payment.is_deleted == False,
            )
            if self.user.role == "income":
                pending_query = pending_query.join(Contract).filter(
                    Contract.sales_person_id == self.user.id
                )
            pending_payments = pending_query.all()

            for p in pending_payments:
                if not p.amount:
                    continue
                diff = abs(float(p.amount) - receipt_amount)
                if diff < 1 or diff / max(float(p.amount), receipt_amount) < 0.05:
                    contract = p.contract
                    if not contract:
                        continue
                    if not self._can_access_contract(contract):
                        continue
                    customer = contract.customer
                    score = 40
                    match_reasons = ["金额匹配"]
                    if receipt_currency and p.currency == receipt_currency:
                        score += 10
                        match_reasons.append("币种匹配")

                    candidates.append({
                        "payment_id": p.id,
                        "contract_id": contract.id,
                        "contract_number": contract.contract_number,
                        "customer_name": customer.name if customer else "未知",
                        "business_type": contract.business_type,
                        "business_description": contract.business_description,
                        "installment_number": p.installment_number,
                        "installment_name": p.installment_name,
                        "amount": float(p.amount) if p.amount else 0,
                        "currency": p.currency,
                        "status": p.status,
                        "paid_date": str(p.paid_date) if p.paid_date else None,
                        "score": score,
                        "match_reason": "、".join(match_reasons),
                    })

        candidates.sort(key=lambda x: x["score"], reverse=True)
        candidates = candidates[:5]

        if not candidates:
            return json.dumps({
                "matches": [],
                "message": "未找到匹配的付款记录。请提供客户姓名以便搜索。",
            }, ensure_ascii=False)

        return json.dumps({
            "matches": candidates,
            "message": f"找到 {len(candidates)} 条可能匹配的付款记录，请确认正确的关联。",
        }, ensure_ascii=False)

    def search_contract_text(self, **kwargs) -> str:
        """按关键词搜索所有合同的全文内容，返回匹配片段"""
        keyword = kwargs.get("keyword", "").strip()
        if not keyword:
            return json.dumps({"error": "缺少 keyword 参数"}, ensure_ascii=False)

        contract_id = kwargs.get("contract_id")

        query = self.db.query(Contract).filter(
            Contract.is_deleted == False,
            Contract.contract_text.isnot(None),
            Contract.contract_text != "",
        )

        if contract_id:
            query = query.filter(Contract.id == contract_id)

        # 角色权限：income 只看自己合同的全文
        if self.user.role == "income":
            query = query.filter(Contract.sales_person_id == self.user.id)

        # ILIKE 模糊搜索
        query = query.filter(Contract.contract_text.ilike(f"%{self._escape_ilike(keyword)}%"))
        contracts = query.limit(10).all()

        if not contracts:
            return json.dumps({
                "matches": [],
                "message": f"未找到包含「{keyword}」的合同内容。",
            }, ensure_ascii=False)

        results = []
        for c in contracts:
            # 提取匹配片段（关键词前后各 80 字）
            text = c.contract_text or ""
            idx = text.lower().find(keyword.lower())
            if idx == -1:
                # fallback：可能是数据库 ILIKE 匹配了但 Python find 大小写不一致
                idx = 0
            start = max(0, idx - 80)
            end = min(len(text), idx + len(keyword) + 80)
            snippet = text[start:end]
            if start > 0:
                snippet = "…" + snippet
            if end < len(text):
                snippet = snippet + "…"

            results.append({
                "contract_id": c.id,
                "contract_number": c.contract_number,
                "customer_name": c.customer.name if c.customer else "未知",
                "business_type": c.business_type,
                "business_description": c.business_description,
                "snippet": snippet,
            })

        return json.dumps({
            "matches": results,
            "message": f"在 {len(results)} 份合同中找到「{keyword}」的匹配内容。",
        }, ensure_ascii=False)

    def ask_contract(self, **kwargs) -> str:
        """检索合同全文内容，返回给 AI 用于回答用户关于合同条款的问题"""
        contract_id = kwargs.get("contract_id")
        question = kwargs.get("question", "").strip()

        if not contract_id:
            return json.dumps({"error": "缺少 contract_id 参数"}, ensure_ascii=False)
        if not question:
            return json.dumps({"error": "缺少 question 参数"}, ensure_ascii=False)

        contract = self.db.query(Contract).filter(
            Contract.id == contract_id,
            Contract.is_deleted == False,
        ).first()

        if not contract:
            return json.dumps({"error": f"合同不存在：{contract_id}"}, ensure_ascii=False)

        if not self._can_access_contract(contract):
            return json.dumps({"error": "无权访问该合同"}, ensure_ascii=False)

        if not contract.contract_text:
            return json.dumps({
                "error": "该合同尚未提取全文内容。请重新上传合同文件以触发全文提取。",
                "contract_number": contract.contract_number,
                "customer_name": contract.customer.name if contract.customer else "未知",
            }, ensure_ascii=False)

        text = contract.contract_text
        if len(text) > 20000:
            text = text[:20000] + "\n…[合同文本过长，已截断]"

        return json.dumps({
            "success": True,
            "contract_id": contract.id,
            "contract_number": contract.contract_number,
            "customer_name": contract.customer.name if contract.customer else "未知",
            "business_type": contract.business_type,
            "signed_date": str(contract.signed_date) if contract.signed_date else None,
            "contract_text": text,
            "user_question": question,
            "_instruction": "请基于以上 contract_text 中的合同全文内容，回答 user_question。只基于原文回答，不要编造任何合同中没有的信息。引用具体条款作为依据。如果合同中没有相关信息，明确告知用户。",
        }, ensure_ascii=False)

    def get_expense_summary(self, **kwargs) -> str:
        """查看支出汇总，按合同或收款方维度聚合"""
        if not self._can_view_expense():
            return json.dumps({"error": "当前角色无权查看支出汇总"}, ensure_ascii=False)

        query = self.db.query(Payment).filter(
            Payment.is_deleted == False,
            Payment.type == "expense",
        )

        if self.user.role == "expense":
            query = query.filter(Payment.created_by == self.user.id)

        if kwargs.get("contract_id"):
            query = query.filter(Payment.contract_id == kwargs["contract_id"])

        if kwargs.get("payee_name"):
            query = query.filter(Payment.payee_name.ilike(f"%{kwargs['payee_name']}%"))

        payments = query.all()
        total_expense = sum(float(p.paid_amount_in_cny or 0) for p in payments)

        group_by = kwargs.get("group_by", "contract")
        groups = {}
        for p in payments:
            if group_by == "payee":
                key = p.payee_name or "未知"
            else:
                key = str(p.contract_id)

            if key not in groups:
                groups[key] = {
                    "total": 0,
                    "count": 0,
                }
                if group_by == "payee":
                    groups[key]["payee_name"] = p.payee_name or "未知"
                else:
                    groups[key]["contract_id"] = p.contract_id
                    groups[key]["contract_number"] = p.contract.contract_number if p.contract else None

            groups[key]["total"] += float(p.paid_amount_in_cny or 0)
            groups[key]["count"] += 1

        return json.dumps({
            "total_expense_in_cny": total_expense,
            "expense_count": len(payments),
            "groups": list(groups.values()),
        }, ensure_ascii=False)

    # ── 分析工具 ──

    def get_payment_summary(self, **kwargs) -> str:
        query = self.db.query(Payment).filter(Payment.is_deleted == False)

        # 角色权限：income 只看收入，expense 只看支出
        if self.user.role == "income":
            query = query.filter(Payment.type == "income")
        elif self.user.role == "expense":
            query = query.filter(Payment.type == "expense")
        elif kwargs.get("type"):
            query = query.filter(Payment.type == kwargs["type"])

        need_contract_join = self.user.role == "income" or kwargs.get("customer_name")
        if need_contract_join:
            query = query.join(Contract)
            if self.user.role == "income":
                query = query.filter(Contract.sales_person_id == self.user.id)
            if kwargs.get("customer_name"):
                query = query.join(Customer).filter(
                    Customer.name.ilike(f"%{kwargs['customer_name']}%")
                )

        if kwargs.get("date_from"):
            try:
                query = query.filter(Payment.paid_date >= date.fromisoformat(kwargs["date_from"]))
            except ValueError:
                pass
        if kwargs.get("date_to"):
            try:
                query = query.filter(Payment.paid_date <= date.fromisoformat(kwargs["date_to"]))
            except ValueError:
                pass

        payments = query.all()

        total_paid = sum(float(p.paid_amount or 0) for p in payments if p.status == "paid")
        total_pending = sum(float(p.amount or 0) for p in payments if p.status == "pending")
        total_overdue = sum(
            float(p.amount or 0)
            for p in payments
            if p.status == "pending" and p.due_date and p.due_date < date.today()
        )

        summary = {
            "total_paid": total_paid,
            "total_pending": total_pending,
            "total_overdue": total_overdue,
            "payment_count": len(payments),
        }

        group_by = kwargs.get("group_by")
        if group_by == "contract":
            groups = {}
            for p in payments:
                cid = p.contract_id
                if cid not in groups:
                    groups[cid] = {
                        "contract_id": cid,
                        "contract_number": p.contract.contract_number if p.contract else None,
                        "paid": 0, "pending": 0,
                    }
                if p.status == "paid":
                    groups[cid]["paid"] += float(p.paid_amount or 0)
                elif p.status == "pending":
                    groups[cid]["pending"] += float(p.amount or 0)
            summary["groups"] = list(groups.values())
        elif group_by == "customer":
            groups = {}
            for p in payments:
                customer_name = p.contract.customer.name if p.contract and p.contract.customer else "未知"
                if customer_name not in groups:
                    groups[customer_name] = {
                        "customer_name": customer_name,
                        "paid": 0, "pending": 0, "contract_count": 0,
                    }
                if p.status == "paid":
                    groups[customer_name]["paid"] += float(p.paid_amount or 0)
                elif p.status == "pending":
                    groups[customer_name]["pending"] += float(p.amount or 0)
            for p in payments:
                customer_name = p.contract.customer.name if p.contract and p.contract.customer else "未知"
                if customer_name in groups:
                    groups[customer_name]["contract_count"] = len(set(
                        pp.contract_id for pp in payments
                        if pp.contract and pp.contract.customer and pp.contract.customer.name == customer_name
                    ))
            summary["groups"] = list(groups.values())

        return json.dumps(summary, ensure_ascii=False)

    def get_overdue_payments(self, **kwargs) -> str:
        query = self.db.query(Payment).filter(
            Payment.is_deleted == False,
            Payment.due_date < date.today(),
            Payment.status == "pending",
        )

        if self.user.role == "income":
            query = query.filter(Payment.type == "income")
            query = query.join(Contract).filter(Contract.sales_person_id == self.user.id)
        elif self.user.role == "expense":
            query = query.filter(Payment.type == "expense")

        if kwargs.get("customer_name"):
            query = query.join(Contract).join(Customer).filter(
                Customer.name.ilike(f"%{kwargs['customer_name']}%")
            )

        min_days = kwargs.get("min_days_overdue", 1)
        payments = query.all()

        results = []
        for p in payments:
            days_overdue = (date.today() - p.due_date).days if p.due_date else 0
            if days_overdue >= min_days:
                results.append({
                    **self._payment_to_dict(p),
                    "days_overdue": days_overdue,
                    "contract_number": p.contract.contract_number if p.contract else None,
                    "customer_name": p.contract.customer.name if p.contract and p.contract.customer else None,
                })

        results.sort(key=lambda x: x["days_overdue"], reverse=True)
        return json.dumps({"overdue_payments": results, "total": len(results)}, ensure_ascii=False)

    def get_expiring_contracts(self, **kwargs) -> str:
        within_days = kwargs.get("within_days", 30)
        target_date = date.today() + timedelta(days=within_days)
        status = kwargs.get("status", "active")

        query = self.db.query(Contract).filter(
            Contract.is_deleted == False,
            Contract.end_date <= target_date,
            Contract.end_date >= date.today(),
        )

        if status:
            query = query.filter(Contract.status == status)

        if self.user.role == "income":
            query = query.filter(Contract.sales_person_id == self.user.id)

        contracts = query.order_by(Contract.end_date).all()
        results = []
        for c in contracts:
            results.append({
                **self._contract_to_dict(c),
                "days_until_expiry": (c.end_date - date.today()).days if c.end_date else None,
            })

        return json.dumps({"contracts": results, "total": len(results)}, ensure_ascii=False)

    # ── 文件分析工具 ──

    @staticmethod
    def _detect_image_mime(header: bytes) -> str:
        """读取文件头判断图片 MIME 类型"""
        if header[:4] == b"\x89PNG":
            return "image/png"
        if header[:3] == b"\xff\xd8\xff":
            return "image/jpeg"
        if header[:4] == b"GIF8":
            return "image/gif"
        if header[:4] == b"RIFF" and header[8:12] == b"WEBP":
            return "image/webp"
        if header[:2] == b"BM":
            return "image/bmp"
        return ""  # 不是已知图片格式

    def _extract_text_sync(self, file_path: str) -> str:
        """同步提取 Word/文本文件内容"""
        # 先尝试 Word
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            if paragraphs:
                return "Word 文档内容:\n" + "\n".join(paragraphs)[:10000]
        except ImportError:
            pass
        except Exception:
            pass

        # 再尝试纯文本（多种编码）
        for encoding in ("utf-8", "gbk", "gb2312", "utf-16"):
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    text = f.read(20000)
                return "文件内容:\n" + text[:10000]
            except (UnicodeDecodeError, UnicodeError):
                continue
        return ""

    def _extract_excel_sync(self, file_path: str) -> str:
        """同步提取 Excel 表格数据"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows.append(f"[工作表: {sheet_name}]")
                count = 0
                for row in ws.iter_rows():
                    vals = [str(c.value) if c.value is not None else "" for c in row]
                    line = "\t".join(vals)
                    if line.strip():
                        rows.append(line)
                        count += 1
                    if count >= 200:
                        rows.append(f"... (仅前 200 行)")
                        break
            wb.close()
            return "Excel 表格数据:\n" + "\n".join(rows)[:10000]
        except ImportError:
            return ""
        except Exception:
            return ".xls 旧格式不受支持，请转换为 .xlsx"

    def analyze_image(self, file_id: str, analysis_type: str = "receipt") -> str:
        """分析上传的文件（图片/PDF/Word/Excel/文本），同步调用"""
        file_path = os.path.join(settings.TEMP_UPLOAD_DIR, file_id)
        if not os.path.exists(file_path):
            return json.dumps({"error": f"文件不存在: {file_id}"}, ensure_ascii=False)

        # 1) 读取文件头判断是否为图片
        with open(file_path, "rb") as f:
            header = f.read(12)
            f.seek(0)
            file_bytes = f.read()

        mime = self._detect_image_mime(header)

        if mime:
            # 是图片 → VL 模型分析
            prompt = {
                "receipt": RECEIPT_ANALYSIS_PROMPT,
                "contract": CONTRACT_ANALYSIS_PROMPT,
                "general": GENERAL_ANALYSIS_PROMPT,
            }.get(analysis_type, GENERAL_ANALYSIS_PROMPT)

            image_base64 = base64.b64encode(file_bytes).decode()
            payload = {
                "model": settings.SILICONFLOW_VISION_MODEL,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{mime};base64,{image_base64}"},
                            },
                            {"type": "text", "text": prompt},
                        ],
                    }
                ],
                "temperature": 0.1,
                "max_tokens": 4096,
            }

            headers = {
                "Authorization": f"Bearer {settings.SILICONFLOW_API_KEY}",
                "Content-Type": "application/json",
            }

            try:
                with httpx.Client(timeout=60.0) as client:
                    response = client.post(
                        f"{settings.SILICONFLOW_BASE_URL}/chat/completions",
                        json=payload,
                        headers=headers,
                    )
                if response.status_code != 200:
                    return json.dumps({"error": f"VL API 错误: {response.text}"}, ensure_ascii=False)

                result = response.json()
                content = result["choices"][0]["message"]["content"]
                try:
                    structured = json.loads(content)
                except json.JSONDecodeError:
                    structured = {"raw": content}

                return json.dumps({
                    "success": True,
                    "data": structured,
                    "file_type": "image",
                    "analysis_type": analysis_type,
                }, ensure_ascii=False)
            except Exception as e:
                logger.exception("analyze_image VL call failed")
                return json.dumps({"error": f"图片分析失败: {str(e)}"}, ensure_ascii=False)

        # 2) PDF 解析
        if header[:4] == b"%PDF":
            try:
                import fitz
                doc = fitz.open(file_path)
                total_pages = len(doc)
                logger.info("PDF分析开始", file_id=file_id, analysis_type=analysis_type, pages=total_pages)

                if analysis_type in ("contract", "receipt"):
                    prompt = {
                        "receipt": RECEIPT_ANALYSIS_PROMPT,
                        "contract": CONTRACT_ANALYSIS_PROMPT,
                    }[analysis_type]

                    # 提取所有页面文本
                    all_page_texts = []
                    for page_num in range(total_pages):
                        text = doc[page_num].get_text()
                        if text.strip():
                            all_page_texts.append(text.strip())
                    doc.close()

                    full_text = "\n\n".join(all_page_texts)

                    if full_text.strip():
                        # ✅ 有文本 → 用 DeepSeek 文本模型解析（快、稳）
                        logger.info("PDF文本提取成功，使用DeepSeek文本模型解析", text_len=len(full_text))
                        try:
                            payload = {
                                "model": settings.DEEPSEEK_AGENT_MODEL,
                                "messages": [{"role": "user", "content": f"{prompt}\n\n以下是合同文件的文字内容，请提取结构化信息：\n\n{full_text[:8000]}"}],
                                "temperature": 0.1,
                                "max_tokens": 4096,
                            }
                            headers = {
                                "Authorization": f"Bearer {settings.DEEPSEEK_API_KEY}",
                                "Content-Type": "application/json",
                            }
                            with httpx.Client(timeout=30.0) as client:
                                response = client.post(
                                    f"{settings.DEEPSEEK_BASE_URL}/chat/completions",
                                    json=payload, headers=headers,
                                )
                            if response.status_code != 200:
                                logger.error("DeepSeek API错误", status=response.status_code, body=response.text[:200])
                                return json.dumps({"error": f"DeepSeek API 错误: {response.text}"}, ensure_ascii=False)

                            content = response.json()["choices"][0]["message"]["content"]
                            try:
                                structured = json.loads(content)
                            except json.JSONDecodeError:
                                structured = {"raw": content}

                            logger.info("PDF文本模型解析完成", analysis_type=analysis_type, keys=list(structured.keys()) if isinstance(structured, dict) else "非dict")
                            return json.dumps({
                                "success": True,
                                "data": structured,
                                "file_type": "pdf",
                                "analysis_type": analysis_type,
                            }, ensure_ascii=False)
                        except Exception as e:
                            logger.exception("DeepSeek解析PDF失败")
                            return json.dumps({"error": f"PDF文本解析失败: {str(e)}"}, ensure_ascii=False)
                    else:
                        # ❌ 无文本（扫描件）→ 渲染为图片走 VL
                        logger.info("PDF无文本（扫描件），使用VL视觉模型分析")
                        try:
                            import fitz as _fitz
                            doc2 = _fitz.open(file_path)
                            pix = doc2[0].get_pixmap(dpi=200)
                            img_bytes = pix.tobytes("png")
                            img_b64 = base64.b64encode(img_bytes).decode()
                            doc2.close()

                            payload = {
                                "model": settings.SILICONFLOW_VISION_MODEL,
                                "messages": [{
                                    "role": "user",
                                    "content": [
                                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
                                        {"type": "text", "text": prompt},
                                    ],
                                }],
                                "temperature": 0.1,
                                "max_tokens": 4096,
                            }
                            headers = {
                                "Authorization": f"Bearer {settings.SILICONFLOW_API_KEY}",
                                "Content-Type": "application/json",
                            }
                            with httpx.Client(timeout=60.0) as client:
                                resp = client.post(f"{settings.SILICONFLOW_BASE_URL}/chat/completions", json=payload, headers=headers)
                            if resp.status_code != 200:
                                return json.dumps({"error": f"VL API 错误: {resp.text}"}, ensure_ascii=False)
                            content = resp.json()["choices"][0]["message"]["content"]
                            try:
                                structured = json.loads(content)
                            except json.JSONDecodeError:
                                structured = {"raw": content}
                            return json.dumps({"success": True, "data": structured, "file_type": "pdf", "analysis_type": analysis_type}, ensure_ascii=False)
                        except Exception as e:
                            logger.exception("PDF扫描件VL分析失败")
                            return json.dumps({"error": f"PDF扫描件分析失败: {str(e)}"}, ensure_ascii=False)

                else:
                    # general 类型：提取文本内容
                    pages_text = []
                    for page_num in range(total_pages):
                        page = doc[page_num]
                        text = page.get_text()
                        if text.strip():
                            pages_text.append(f"第{page_num + 1}页:\n{text.strip()}")
                        else:
                            pix = page.get_pixmap(dpi=200)
                            img_bytes = pix.tobytes("png")
                            img_b64 = base64.b64encode(img_bytes).decode()
                            payload = {
                                "model": settings.SILICONFLOW_VISION_MODEL,
                                "messages": [{
                                    "role": "user",
                                    "content": [
                                        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
                                        {"type": "text", "text": GENERAL_ANALYSIS_PROMPT},
                                    ],
                                }],
                                "temperature": 0.1,
                                "max_tokens": 4096,
                            }
                            headers = {
                                "Authorization": f"Bearer {settings.SILICONFLOW_API_KEY}",
                                "Content-Type": "application/json",
                            }
                            with httpx.Client(timeout=60.0) as client:
                                resp = client.post(
                                    f"{settings.SILICONFLOW_BASE_URL}/chat/completions",
                                    json=payload, headers=headers,
                                )
                            if resp.status_code == 200:
                                content = resp.json()["choices"][0]["message"]["content"]
                                pages_text.append(f"第{page_num + 1}页(扫描): {content[:500]}")
                            else:
                                pages_text.append(f"第{page_num + 1}页分析失败")
                    doc.close()
                    result = "\n".join(pages_text)
                    return json.dumps({"success": True, "data": {"content": result[:5000]}, "file_type": "pdf"}, ensure_ascii=False)
            except ImportError:
                return json.dumps({"error": "PDF 分析不可用（缺少 PyMuPDF）"}, ensure_ascii=False)
            except Exception as e:
                logger.exception("PDF分析异常")
                return json.dumps({"error": f"PDF 分析失败: {str(e)}"}, ensure_ascii=False)

        # 3) 尝试 Word / 文本
        text_result = self._extract_text_sync(file_path)
        if text_result:
            return json.dumps({"success": True, "data": {"content": text_result}, "file_type": "document"}, ensure_ascii=False)

        # 4) 尝试 Excel
        excel_result = self._extract_excel_sync(file_path)
        if excel_result:
            return json.dumps({"success": True, "data": {"content": excel_result}, "file_type": "excel"}, ensure_ascii=False)

        return json.dumps({"error": "无法识别的文件类型，支持：图片、PDF、Word、Excel、文本"}, ensure_ascii=False)

    def execute(self, tool_name: str, arguments: dict) -> str:
        """统一执行入口"""
        handler = getattr(self, tool_name, None)
        if not handler:
            logger.warning("未知工具调用: %s", tool_name)
            return json.dumps({"error": f"未知工具: {tool_name}"}, ensure_ascii=False)

        args_preview = json.dumps(arguments, ensure_ascii=False, default=str)[:300]
        logger.info("🔧 工具调用: %s | 参数: %s", tool_name, args_preview)

        try:
            result = handler(**arguments)
            logger.info("✅ 工具结果: %s → %s", tool_name, result[:200] if result else "empty")
            return result
        except Exception as e:
            logger.exception("❌ 工具执行失败: %s", tool_name)
            return json.dumps({"error": f"工具执行失败: {str(e)}"}, ensure_ascii=False)


# OpenAI function calling 格式的工具定义
TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "search_customers",
            "description": "按姓名、电话或微信群名搜索客户（自动兼容繁体/简体）。返回匹配的客户列表及其关联合同数量。",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "客户姓名（支持模糊匹配）"},
                    "phone": {"type": "string", "description": "电话号码"},
                    "wechat_group": {"type": "string", "description": "微信群名称"},
                    "limit": {"type": "integer", "description": "最大返回数量，默认10", "default": 10},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_customer",
            "description": "创建客户记录。如果同名+同电话/邮箱的客户已存在，会返回已有客户（不会重复创建）。从合同文件提取到客户信息后应调用此工具。返回包含 customer.id 用于后续创建合同。",
            "parameters": {
                "type": "object",
                "required": ["name"],
                "properties": {
                    "name": {"type": "string", "description": "客户姓名（必填）"},
                    "phone": {"type": "string", "description": "联系电话（phone和email至少填一个）"},
                    "email": {"type": "string", "description": "联系邮箱（phone和email至少填一个）"},
                    "contact_person": {"type": "string", "description": "联系人"},
                    "id_card_number": {"type": "string", "description": "身份证号或证件号"},
                    "wechat_group_name": {"type": "string", "description": "微信群名称"},
                    "address": {"type": "string", "description": "地址"},
                    "remarks": {"type": "string", "description": "备注"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_customer",
            "description": "更新已有客户信息。当从合同文件中提取到客户电话、证件号等新信息时，用此工具补充到已有客户记录中。",
            "parameters": {
                "type": "object",
                "required": ["customer_id"],
                "properties": {
                    "customer_id": {"type": "integer", "description": "客户ID"},
                    "phone": {"type": "string", "description": "联系电话"},
                    "email": {"type": "string", "description": "联系邮箱"},
                    "id_card_number": {"type": "string", "description": "身份证号或证件号"},
                    "wechat_group_name": {"type": "string", "description": "微信群名称"},
                    "address": {"type": "string", "description": "地址"},
                    "remarks": {"type": "string", "description": "备注"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_contracts",
            "description": "按合同编号、客户名、状态或关键词搜索合同。返回合同列表含金额和付款进度。",
            "parameters": {
                "type": "object",
                "properties": {
                    "contract_number": {"type": "string", "description": "合同编号"},
                    "customer_name": {"type": "string", "description": "客户姓名（模糊匹配）"},
                    "status": {
                        "type": "string",
                        "enum": ["active", "completed"],
                        "description": "合同状态筛选",
                    },
                    "keyword": {"type": "string", "description": "全文搜索关键词"},
                    "date_from": {"type": "string", "description": "签订日期起始（YYYY-MM-DD）"},
                    "date_to": {"type": "string", "description": "签订日期截止（YYYY-MM-DD）"},
                    "page": {"type": "integer", "default": 1},
                    "per_page": {"type": "integer", "default": 10},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_contract_detail",
            "description": "获取合同完整详情，包含所有付款记录和付款进度。用于用户询问某个具体合同时。",
            "parameters": {
                "type": "object",
                "required": ["contract_id"],
                "properties": {
                    "contract_id": {"type": "integer", "description": "合同ID"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_customer_contracts",
            "description": "获取某客户的所有合同及付款状态汇总。",
            "parameters": {
                "type": "object",
                "required": ["customer_id"],
                "properties": {
                    "customer_id": {"type": "integer", "description": "客户ID"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "query_payments",
            "description": "按合同ID、类型或状态查询付款记录。可查找逾期付款。income角色只能查看收入，expense角色只能查看支出。",
            "parameters": {
                "type": "object",
                "properties": {
                    "contract_id": {"type": "integer", "description": "按合同ID筛选"},
                    "type": {
                        "type": "string",
                        "enum": ["income", "expense"],
                        "description": "付款类型筛选：income（收入）或 expense（支出）",
                    },
                    "status": {
                        "type": "string",
                        "enum": ["pending", "paid"],
                        "description": "付款状态筛选",
                    },
                    "overdue_only": {"type": "boolean", "description": "只返回逾期付款", "default": False},
                    "page": {"type": "integer", "default": 1},
                    "per_page": {"type": "integer", "default": 20},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_contract",
            "description": "为客户创建合同记录。需要先通过 create_customer 或 search_customers 获取 customer_id。合同编号自动生成。如果同一文件已创建过合同会返回已有记录。创建后系统会自动根据合同中的付款条款创建对应的付款记录（已付款项自动记录为 pending，如同时上传了凭证且金额匹配则直接记录为 paid）。",
            "parameters": {
                "type": "object",
                "required": ["customer_id", "file_id", "contract_data"],
                "properties": {
                    "customer_id": {"type": "integer", "description": "客户ID（通过 create_customer 或 search_customers 获取）"},
                    "file_id": {"type": "string", "description": "上传文件的ID（聊天上传时返回的 file_id）"},
                    "contract_data": {"type": "object", "description": "从合同文件提取的全部信息（JSON对象，包含甲方乙方、金额、payment_terms等）。payment_terms中每个款项必须包含is_paid布尔字段，表示该款项是否已付"},
                    "title": {"type": "string", "description": "合同标题（如：购车合同、两地牌办理合同）"},
                    "total_amount": {"type": "number", "description": "合同总金额"},
                    "currency": {"type": "string", "enum": ["CNY", "HKD", "USD"], "description": "币种，默认CNY"},
                    "signed_date": {"type": "string", "description": "签订日期（YYYY-MM-DD）"},
                    "business_type": {"type": "string", "enum": ["车辆业务", "中港牌业务"], "description": "业务大类：车辆业务（买车卖车）或中港牌业务（办理中港车牌）"},
                    "business_description": {"type": "string", "description": "业务具体描述，如：购买丰田阿尔法30系、办理深圳湾口岸中港车牌"},
                    "wechat_group": {"type": "string", "description": "业务微信群名称（如有）"},
                    "receipt_data": {"type": "object", "description": "如果同时上传了付款凭证，传入凭证分析结果（JSON对象）。系统会自动匹配合同中已付款项"},
                    "receipt_file_path": {"type": "string", "description": "如果同时上传了付款凭证图片，传入凭证文件路径"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_contract",
            "description": "更新合同信息。用于补充微信群名称、备注等信息。当用户发送业务群截图时，提取群名后用此工具关联到合同。",
            "parameters": {
                "type": "object",
                "required": ["contract_id"],
                "properties": {
                    "contract_id": {"type": "integer", "description": "合同ID"},
                    "wechat_group": {"type": "string", "description": "业务微信群名称"},
                    "remarks": {"type": "string", "description": "备注"},
                    "title": {"type": "string", "description": "合同标题"},
                    "business_description": {"type": "string", "description": "业务描述"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_payment",
            "description": "为合同创建收入付款记录（客户向公司付款）。无凭证时创建为 pending 状态（不参与结算），有凭证时创建为 paid 状态（自动参与结算）。自动按付款日期查找汇率并折算为人民币。仅用于手动录入或凭证上传场景——合同录入时系统会自动创建已付款项的记录，不需要手动调用。",
            "parameters": {
                "type": "object",
                "required": ["contract_id", "installment_number", "amount", "currency", "paid_date"],
                "properties": {
                    "contract_id": {"type": "integer", "description": "合同ID"},
                    "installment_number": {"type": "integer", "description": "第几期（1, 2, 3...）"},
                    "installment_name": {"type": "string", "description": "期数名称（如「定金」、「首期」、「尾款」），从合同原文提取"},
                    "amount": {"type": "number", "description": "付款金额"},
                    "currency": {"type": "string", "enum": ["CNY", "HKD", "USD"], "default": "CNY"},
                    "paid_date": {"type": "string", "description": "实际付款日期（YYYY-MM-DD）"},
                    "payment_method": {
                        "type": "string",
                        "enum": ["bank_transfer", "wechat", "alipay", "cash", "check", "unknown"],
                        "description": "付款方式",
                    },
                    "receipt_image_path": {"type": "string", "description": "付款凭证图片路径（有凭证时传入）"},
                    "notes": {"type": "string", "description": "备注"},
                    "receipt_data": {"type": "object", "description": "凭证分析结构化数据（JSON对象，包含document_type/amount/payer_name/transaction_id等）"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "create_expense",
            "description": "为合同创建支出记录（公司向第三方付款，如渠道费、办证费等）。需要指定收款方名称。期数自动生成。无凭证时创建为 pending 状态（不参与结算），有凭证时创建为 paid 状态（自动参与结算）。自动按付款日期查找汇率并折算为人民币。仅admin和expense角色可用。",
            "parameters": {
                "type": "object",
                "required": ["contract_id", "amount", "currency", "paid_date", "payee_name"],
                "properties": {
                    "contract_id": {"type": "integer", "description": "合同ID"},
                    "amount": {"type": "number", "description": "支出金额"},
                    "currency": {"type": "string", "enum": ["CNY", "HKD", "USD"], "default": "CNY"},
                    "paid_date": {"type": "string", "description": "实际付款日期（YYYY-MM-DD）"},
                    "payee_name": {"type": "string", "description": "收款方名称（如：某某代办公司）"},
                    "installment_name": {"type": "string", "description": "期数名称（如「渠道费」、「办证费」）"},
                    "payment_method": {
                        "type": "string",
                        "enum": ["bank_transfer", "wechat", "alipay", "cash", "check", "unknown"],
                        "description": "付款方式",
                    },
                    "receipt_image_path": {"type": "string", "description": "付款凭证图片路径（有凭证时传入）"},
                    "notes": {"type": "string", "description": "备注"},
                    "receipt_data": {"type": "object", "description": "凭证分析结构化数据（JSON对象）"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "update_payment",
            "description": "更新已有付款记录的备注、凭证、付款方式等信息。当用户为已有付款补充凭证时使用此工具（而不是创建新记录）。补充凭证后系统自动将 pending 转为 paid 并参与结算。",
            "parameters": {
                "type": "object",
                "required": ["payment_id"],
                "properties": {
                    "payment_id": {"type": "integer", "description": "付款记录ID"},
                    "notes": {"type": "string", "description": "备注信息（根据凭证内容自动生成描述性备注）"},
                    "payment_method": {"type": "string", "description": "付款方式"},
                    "receipt_image_path": {"type": "string", "description": "凭证图片路径"},
                    "receipt_data": {"type": "object", "description": "凭证分析结构化数据（JSON对象）"},
                    "installment_name": {"type": "string", "description": "期数名称"},
                    "paid_date": {"type": "string", "description": "实际付款日期（YYYY-MM-DD）"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "match_receipt",
            "description": "根据凭证分析结果智能匹配到合同付款记录。优先按客户名匹配，其次按金额匹配。返回候选列表供用户确认，确认后用 update_payment 补充凭证。用户上传付款凭证时必须先调用此工具查找匹配。",
            "parameters": {
                "type": "object",
                "required": ["receipt_data"],
                "properties": {
                    "receipt_data": {"type": "object", "description": "凭证分析结果（JSON对象，包含 payer_name/amount/currency/transaction_date 等）"},
                    "customer_name": {"type": "string", "description": "客户姓名（当凭证中无法识别客户时，由用户提供）"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_expense_summary",
            "description": "查看支出汇总，可按合同或收款方维度聚合。仅admin和expense角色可用。",
            "parameters": {
                "type": "object",
                "properties": {
                    "contract_id": {"type": "integer", "description": "按合同ID筛选"},
                    "payee_name": {"type": "string", "description": "按收款方名称筛选（模糊匹配）"},
                    "group_by": {
                        "type": "string",
                        "enum": ["contract", "payee"],
                        "description": "分组方式：按合同或按收款方",
                        "default": "contract",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_payment_summary",
            "description": "获取付款汇总：已付总额、待付总额、逾期总额。可按客户或合同分组。",
            "parameters": {
                "type": "object",
                "properties": {
                    "customer_name": {"type": "string", "description": "按客户名筛选"},
                    "date_from": {"type": "string", "description": "日期范围起始（YYYY-MM-DD）"},
                    "date_to": {"type": "string", "description": "日期范围截止（YYYY-MM-DD）"},
                    "group_by": {
                        "type": "string",
                        "enum": ["customer", "contract", "month"],
                        "description": "分组方式",
                    },
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_overdue_payments",
            "description": "查找所有逾期未付的款项（已过应付款日期且状态为pending）。",
            "parameters": {
                "type": "object",
                "properties": {
                    "customer_name": {"type": "string", "description": "按客户名筛选"},
                    "min_days_overdue": {"type": "integer", "description": "最小逾期天数", "default": 1},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "get_expiring_contracts",
            "description": "查找即将到期的合同（end_date在指定天数内）。",
            "parameters": {
                "type": "object",
                "properties": {
                    "within_days": {"type": "integer", "description": "从今天起的天数", "default": 30},
                    "status": {"type": "string", "default": "active"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_contract_text",
            "description": "按关键词搜索所有合同的全文内容，返回匹配的合同列表和文本片段。用于查找包含特定条款、约定的合同（如搜索'违约金'、'仲裁'、'交车日期'等）。",
            "parameters": {
                "type": "object",
                "required": ["keyword"],
                "properties": {
                    "keyword": {"type": "string", "description": "搜索关键词，如'违约金'、'仲裁'、'交车日期'等"},
                    "contract_id": {"type": "integer", "description": "限定在某份合同中搜索（可选）"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "ask_contract",
            "description": "检索某份合同的全文内容，用于回答关于合同具体条款的问题（如违约责任、付款条件、交车时间、双方权利义务等）。调用后系统会基于合同原文回答用户问题。",
            "parameters": {
                "type": "object",
                "required": ["contract_id", "question"],
                "properties": {
                    "contract_id": {"type": "integer", "description": "合同ID"},
                    "question": {"type": "string", "description": "用户关于合同条款的具体问题，如'违约金怎么约定'、'交车截止日期是什么'、'甲方有什么义务'"},
                },
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "analyze_image",
            "description": "分析上传的文件（支持图片、PDF、Word、Excel、文本文件）。图片/扫描件通过视觉模型提取信息，文档类提取文字内容。文件需先通过聊天上传接口上传。",
            "parameters": {
                "type": "object",
                "required": ["file_id", "analysis_type"],
                "properties": {
                    "file_id": {"type": "string", "description": "上传接口返回的文件ID"},
                    "analysis_type": {
                        "type": "string",
                        "enum": ["receipt", "contract", "general"],
                        "description": "分析类型（图片/扫描件适用）",
                    },
                },
            },
        },
    },
]
