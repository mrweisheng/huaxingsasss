"""
ContractAgent — 基于 ReAct 模式的智能业务助手
使用 DeepSeek API 进行函数调用，SiliconFlow VL 模型处理图片
"""
import json
import logging
import os
import uuid
from datetime import datetime, date
from typing import AsyncGenerator, Optional, List, Dict, Any

from sqlalchemy.orm import Session

from app.ai.llm_client import DeepSeekClient
from app.ai.prompts import (
    build_system_prompt,
    RECEIPT_ANALYSIS_PROMPT,
    CONTRACT_ANALYSIS_PROMPT,
    SUMMARY_PROMPT,
)
from app.ai.tools import ToolExecutor, TOOL_DEFINITIONS
from app.config import settings
from app.models.chat_history import ChatHistory
from app.models.user import User

logger = logging.getLogger(__name__)


class ContractAgent:
    def __init__(self, db: Session, user: User):
        self.db = db
        self.user = user
        self.llm = DeepSeekClient()
        self.executor = ToolExecutor(db, user)

    async def chat(
        self,
        session_id: Optional[str],
        user_message: str,
        attachments: Optional[List[dict]] = None,
    ) -> AsyncGenerator[dict, None]:
        """
        主对话方法，SSE 流式输出。

        Yields:
            {"event": "text", "data": {"content": "..."}}
            {"event": "tool_call", "data": {"name": "...", "arguments": ...}}
            {"event": "tool_result", "data": {"name": "...", "result": "..."}}
            {"event": "thinking", "data": {"message": "..."}}
            {"event": "done", "data": {"session_id": "...", "tokens_used": ...}}
        """
        # 创建或复用会话
        if not session_id:
            session_id = str(uuid.uuid4())

        # 保存用户消息
        self._save_message(session_id, "user", user_message, metadata={"attachments": attachments})

        # 处理附件（图片/PDF/文档分析）
        file_context = ""
        if attachments:
            file_context = await self._process_attachments(attachments)
            yield {"event": "thinking", "data": {"message": "文件分析完成"}}

        # 加载对话历史
        history = self._load_history(session_id)

        # 构建消息
        messages = self._build_messages(history, user_message, file_context)

        # ReAct 循环
        total_tokens = 0
        for iteration in range(settings.AGENT_MAX_ITERATIONS):
            # 调用 LLM
            full_text = ""
            tool_calls: List[dict] = []

            async for event in self.llm.chat_completion_stream(
                messages=messages,
                tools=TOOL_DEFINITIONS,
            ):
                if event["type"] == "text":
                    full_text += event["content"]
                    yield {"event": "text", "data": {"content": event["content"]}}

                elif event["type"] == "tool_call":
                    tool_calls.append({
                        "id": event["id"],
                        "name": event["name"],
                        "arguments": event["arguments"],
                    })

                elif event["type"] == "usage":
                    total_tokens += event.get("total_tokens", 0)

                elif event["type"] == "done":
                    if event.get("finish_reason") == "tool_calls":
                        pass

            if not tool_calls:
                # 无工具调用，对话结束
                self._save_message(session_id, "assistant", full_text)
                yield {
                    "event": "done",
                    "data": {
                        "session_id": session_id,
                        "tokens_used": total_tokens,
                    },
                }
                return

            # 执行工具调用
            # 先把助手的工具调用消息加入上下文
            assistant_msg = {"role": "assistant", "content": full_text or None}
            assistant_msg["tool_calls"] = [
                {
                    "id": tc["id"],
                    "type": "function",
                    "function": {
                        "name": tc["name"],
                        "arguments": tc["arguments"],
                    },
                }
                for tc in tool_calls
            ]
            messages.append(assistant_msg)

            for tc in tool_calls:
                yield {
                    "event": "tool_call",
                    "data": {
                        "name": tc["name"],
                        "arguments": tc["arguments"],
                    },
                }

                # 执行工具
                try:
                    args = json.loads(tc["arguments"]) if isinstance(tc["arguments"], str) else tc["arguments"]
                except json.JSONDecodeError:
                    args = {}

                result = self.executor.execute(tc["name"], args)

                yield {
                    "event": "tool_result",
                    "data": {"name": tc["name"], "result": result[:2000]},
                }

                # 工具结果加入上下文
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc["id"],
                    "content": result,
                })

            # 继续循环，让 LLM 根据工具结果生成回复

        # 超过最大迭代次数
        self._save_message(session_id, "assistant", full_text)
        yield {
            "event": "done",
            "data": {"session_id": session_id, "tokens_used": total_tokens},
        }

    async def _process_attachments(self, attachments: List[dict]) -> str:
        """处理附件（图片/PDF/Word/Excel/文本），返回分析结果的文本描述"""
        results = []
        for att in attachments:
            file_id = att.get("file_id", "")
            file_type = att.get("file_type", "image")

            file_path = os.path.join(settings.TEMP_UPLOAD_DIR, file_id)
            if not os.path.exists(file_path):
                results.append(f"文件 {file_id} 不存在")
                continue

            if file_type == "pdf":
                result = await self._process_pdf(file_path)
                results.append(result)
            elif file_type == "word":
                text = self._extract_text(file_path, "word")
                results.append(f"Word 文档内容:\n{text}")
            elif file_type == "excel":
                text = self._extract_excel(file_path)
                results.append(f"Excel 表格数据:\n{text}")
            elif file_type == "text":
                text = self._extract_text(file_path, "text")
                results.append(f"文本文件内容:\n{text}")
            else:
                # 图片处理（原有逻辑）
                prompt = RECEIPT_ANALYSIS_PROMPT if file_type == "receipt" else CONTRACT_ANALYSIS_PROMPT
                try:
                    analysis = await self.llm.analyze_image_with_vl(file_path, prompt)
                    results.append(f"图片分析结果: {json.dumps(analysis.get('data', {}), ensure_ascii=False)}")
                except Exception as e:
                    logger.exception("Image analysis failed")
                    results.append(f"图片分析失败: {str(e)}")

        return "\n".join(results)

    def _extract_text(self, file_path: str, file_type: str) -> str:
        """提取 Word 文档或纯文本文件内容"""
        if file_type == "word":
            try:
                from docx import Document
                doc = Document(file_path)
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                if not paragraphs:
                    return "Word 文档内容为空或为纯图片文档"
                text = "\n".join(paragraphs)
                # 限制长度避免超出 token 限制
                return text[:10000] if len(text) > 10000 else text
            except ImportError:
                return "Word 文档解析不可用（缺少 python-docx 依赖）"
            except Exception as e:
                return f"Word 文档解析失败: {str(e)}"

        # 纯文本文件
        try:
            for encoding in ("utf-8", "gbk", "gb2312", "utf-16"):
                try:
                    with open(file_path, "r", encoding=encoding) as f:
                        text = f.read(20000)
                    return text[:10000] if len(text) > 10000 else text
                except (UnicodeDecodeError, UnicodeError):
                    continue
            return "文本文件无法解码（尝试了 utf-8、gbk、gb2312、utf-16）"
        except Exception as e:
            return f"文本文件读取失败: {str(e)}"

    def _extract_excel(self, file_path: str) -> str:
        """提取 Excel 表格数据"""
        try:
            import openpyxl
        except ImportError:
            return "Excel 解析不可用（缺少 openpyxl 依赖）"

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            # .xls 旧格式不支持
            return "无法解析该 Excel 文件（仅支持 .xlsx 格式，旧版 .xls 请转换为 .xlsx 后重试）"

        rows_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text.append(f"[工作表: {sheet_name}]")
            count = 0
            for row in ws.iter_rows():
                vals = [str(c.value) if c.value is not None else "" for c in row]
                line = "\t".join(vals)
                if line.strip():
                    rows_text.append(line)
                    count += 1
                if count >= 200:  # 最多 200 行
                    rows_text.append(f"... (仅显示前 200 行，共 {ws.max_row} 行)")
                    break

        wb.close()
        text = "\n".join(rows_text)
        return text[:10000] if len(text) > 10000 else text

    async def _process_pdf(self, file_path: str) -> str:
        """处理 PDF 文件：先尝试提取文本，文本量少则渲染页面为图片送 VL 分析"""
        try:
            import fitz
        except ImportError:
            return "PDF 分析功能不可用（缺少 PyMuPDF 依赖）"

        try:
            doc = fitz.open(file_path)
        except Exception:
            return f"PDF 文件无法解析"

        num_pages = len(doc)

        # Step 1: 尝试提取文本
        all_text = ""
        for page_num in range(num_pages):
            page = doc[page_num]
            text = page.get_text()
            if text.strip():
                all_text += f"\n--- 第{page_num + 1}/{num_pages}页 ---\n{text.strip()}"

        # 文本量充足（>100字符），直接返回文本
        if len(all_text.strip()) >= 100:
            doc.close()
            return f"PDF 文本提取成功（共{num_pages}页）：{all_text}"

        # Step 2: 文本量不足 → 判为扫描件，逐页渲染为图片送 VL 分析
        logger.info("PDF text too short (%d chars), rendering pages as images", len(all_text.strip()))

        page_results = []
        for page_num in range(num_pages):
            page = doc[page_num]
            pix = page.get_pixmap(dpi=200)
            img_bytes = pix.tobytes("png")

            tmp_path = os.path.join(settings.TEMP_UPLOAD_DIR, f"{uuid.uuid4()}.png")
            try:
                with open(tmp_path, "wb") as f:
                    f.write(img_bytes)

                analysis = await self.llm.analyze_image_with_vl(tmp_path, CONTRACT_ANALYSIS_PROMPT)
                page_results.append(
                    f"第{page_num + 1}/{num_pages}页: {json.dumps(analysis.get('data', {}), ensure_ascii=False)}"
                )
            except Exception as e:
                logger.exception("PDF page %d analysis failed", page_num)
                page_results.append(f"第{page_num + 1}/{num_pages}页分析失败: {str(e)}")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        doc.close()

        if len(page_results) == 1:
            return page_results[0]
        return f"PDF 共有 {num_pages} 页，逐页分析结果：\n" + "\n".join(page_results)

    def _load_history(self, session_id: str) -> List[dict]:
        """从数据库加载对话历史"""
        records = (
            self.db.query(ChatHistory)
            .filter(
                ChatHistory.session_id == session_id,
                ChatHistory.user_id == self.user.id,
            )
            .order_by(ChatHistory.created_at.desc())
            .limit(settings.AGENT_HISTORY_WINDOW)
            .all()
        )
        records.reverse()

        messages = []
        for r in records:
            if r.role == "user":
                content = r.question
                meta = r.extra_metadata or {}
                if meta.get("file_context"):
                    content += f"\n\n[文件分析上下文]\n{meta['file_context']}"
                messages.append({"role": "user", "content": content})
            elif r.role == "assistant":
                msg = {"role": "assistant", "content": r.answer}
                if r.tool_calls:
                    msg["tool_calls"] = r.tool_calls
                messages.append(msg)
            elif r.role == "tool":
                messages.append({
                    "role": "tool",
                    "tool_call_id": r.extra_metadata.get("tool_call_id", "") if r.extra_metadata else "",
                    "content": r.answer or "",
                })

        return messages

    def _build_messages(
        self,
        history: List[dict],
        user_message: str,
        file_context: str = "",
    ) -> List[dict]:
        """构建完整的消息数组"""
        system = {
            "role": "system",
            "content": build_system_prompt(
                user_name=self.user.full_name or self.user.username,
                user_role=self.user.role,
                current_date=date.today().isoformat(),
            ),
        }

        messages = [system]

        # 历史（不包含当前用户消息，因为用户消息已经在最后一条）
        messages.extend(history)

        # 如果历史为空或最后一条不是当前消息，添加当前消息
        if file_context:
            full_content = f"{user_message}\n\n[文件分析上下文]\n{file_context}"
        else:
            full_content = user_message

        if not history or history[-1].get("role") != "user":
            messages.append({"role": "user", "content": full_content})

        return messages

    def _save_message(
        self,
        session_id: str,
        role: str,
        content: str,
        tool_calls: Optional[list] = None,
        metadata: Optional[dict] = None,
    ):
        """保存消息到数据库"""
        if role == "user":
            question = content
            answer = None
        else:
            question = ""
            answer = content

        record = ChatHistory(
            user_id=self.user.id,
            session_id=session_id,
            question=question,
            answer=answer,
            role=role,
            tool_calls=tool_calls,
            extra_metadata=metadata or {},
            llm_model=settings.DEEPSEEK_AGENT_MODEL,
        )
        self.db.add(record)
        self.db.commit()

    def save_tool_message(
        self,
        session_id: str,
        tool_call_id: str,
        tool_name: str,
        result: str,
    ):
        """保存工具调用结果消息"""
        record = ChatHistory(
            user_id=self.user.id,
            session_id=session_id,
            question="",
            answer=result,
            role="tool",
            intent_type=tool_name,
            extra_metadata={"tool_call_id": tool_call_id},
            llm_model=settings.DEEPSEEK_AGENT_MODEL,
        )
        self.db.add(record)
        self.db.commit()

    def get_sessions(self) -> List[dict]:
        """获取用户的所有会话列表（2 次查询替代 N+1）"""
        from sqlalchemy import func as sa_func

        # Query 1: 聚合查询 — 每会话的消息数和最后活动时间
        aggregates = (
            self.db.query(
                ChatHistory.session_id,
                sa_func.count(ChatHistory.id).label("message_count"),
                sa_func.max(ChatHistory.created_at).label("last_activity"),
            )
            .filter(ChatHistory.user_id == self.user.id)
            .group_by(ChatHistory.session_id)
            .all()
        )

        session_ids = [s.session_id for s in aggregates]

        # Query 2: 取每会话首条用户消息作标题（使用窗口函数）
        first_msg_map: dict[str, Optional[str]] = {}
        if session_ids:
            subq = (
                self.db.query(
                    ChatHistory.session_id,
                    ChatHistory.question,
                    sa_func.row_number().over(
                        partition_by=ChatHistory.session_id,
                        order_by=ChatHistory.created_at,
                    ).label("rn"),
                )
                .filter(
                    ChatHistory.user_id == self.user.id,
                    ChatHistory.role == "user",
                    ChatHistory.session_id.in_(session_ids),
                )
                .subquery()
            )
            first_msgs = (
                self.db.query(subq.c.session_id, subq.c.question)
                .filter(subq.c.rn == 1)
                .all()
            )
            for sid, question in first_msgs:
                first_msg_map[sid] = question[:50] if question else None

        result = []
        for s in aggregates:
            result.append({
                "session_id": s.session_id,
                "created_at": s.last_activity.isoformat() if s.last_activity else None,
                "message_count": s.message_count,
                "title": first_msg_map.get(s.session_id),
            })

        result.sort(key=lambda x: x["created_at"] or "", reverse=True)
        return result

    def get_history(self, session_id: str) -> List[dict]:
        """获取会话历史"""
        records = (
            self.db.query(ChatHistory)
            .filter(
                ChatHistory.session_id == session_id,
                ChatHistory.user_id == self.user.id,
            )
            .order_by(ChatHistory.created_at.asc())
            .all()
        )

        return [
            {
                "id": r.id,
                "role": r.role,
                "content": r.question if r.role == "user" else r.answer,
                "tool_calls": r.tool_calls,
                "intent_type": r.intent_type,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in records
        ]

    def delete_session(self, session_id: str) -> bool:
        """删除会话及其所有消息"""
        deleted = (
            self.db.query(ChatHistory)
            .filter(
                ChatHistory.session_id == session_id,
                ChatHistory.user_id == self.user.id,
            )
            .delete()
        )
        self.db.commit()
        return deleted > 0
