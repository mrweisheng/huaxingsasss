"""
共享文件分析工具函数
从 contract_analyzer.py 提取，供 ContractAnalyzer 和 ReceiptAnalyzer 共同引用。
"""
import base64
import io
import json
import logging
import os

import httpx

from app.config import settings

# 注册 HEIF/HEIC 解码器到 Pillow，让 Image.open() 直接识别 .heic/.heif
# 副作用安全：register 多次等价于 register 一次
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
except ImportError:
    # 缺依赖时降级：HEIC 上传会在 upload 端点转码阶段抛 400，下游不受影响
    pass

logger = logging.getLogger(__name__)

# 图片压缩参数
MAX_IMAGE_DIMENSION = 1600
JPEG_QUALITY = 85


def compress_image(file_bytes: bytes, mime: str) -> tuple:
    """压缩图片：缩放到 MAX_IMAGE_DIMENSION 内 + JPEG 质量 85。"""
    from PIL import Image

    try:
        img = Image.open(io.BytesIO(file_bytes))
    except Exception:
        return file_bytes, mime

    w, h = img.size
    if max(w, h) <= MAX_IMAGE_DIMENSION and mime == "image/jpeg" and len(file_bytes) < 500_000:
        return file_bytes, mime

    if max(w, h) > MAX_IMAGE_DIMENSION:
        ratio = MAX_IMAGE_DIMENSION / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

    buf = io.BytesIO()
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    img.save(buf, format="JPEG", quality=JPEG_QUALITY)

    compressed = buf.getvalue()
    logger.info("图片压缩: %dx%d %s %.0fKB -> %.0fKB", w, h, mime, len(file_bytes)/1024, len(compressed)/1024)
    return compressed, "image/jpeg"


def detect_image_mime(header: bytes) -> str:
    """读取文件头判断图片 MIME 类型，未识别返回空字符串。"""
    if header[:4] == b"\x89PNG":
        return "image/png"
    if header[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if header[:4] == b"GIF8":
        return "image/gif"
    if header[:4] == b"RIFF" and len(header) > 11 and header[8:12] == b"WEBP":
        return "image/webp"
    if header[:2] == b"BM":
        return "image/bmp"
    # HEIC/HEIF（ISOBMFF "ftyp" box + brand 子类型，iPhone 默认拍照格式）
    if len(header) >= 12 and header[4:8] == b"ftyp" and header[8:12] in (
        b"heic", b"heix", b"hevc", b"heim", b"heis", b"hevm", b"hevs", b"mif1", b"msf1"
    ):
        return "image/heic"
    return ""


def guess_extension(content: bytes) -> str:
    """通过文件头判断扩展名"""
    if content[:4] == b"%PDF":
        return ".pdf"
    if content[:3] == b"\xff\xd8\xff":
        return ".jpg"
    if content[:4] == b"\x89PNG":
        return ".png"
    if content[:4] == b"GIF8":
        return ".gif"
    if content[:4] == b"RIFF" and len(content) > 11 and content[8:12] == b"WEBP":
        return ".webp"
    # HEIC/HEIF（iPhone 默认拍照格式）
    if len(content) >= 12 and content[4:8] == b"ftyp" and content[8:12] in (
        b"heic", b"heix", b"hevc", b"heim", b"heis", b"hevm", b"hevs", b"mif1", b"msf1"
    ):
        return ".heic"
    # Office Open XML (ZIP-based): .docx / .xlsx / .pptx 都以 PK 开头
    if content[:4] == b"PK\x03\x04":
        # 进一步区分 Word / Excel
        # ZIP 内文件列表中通常包含对应的 content type 标识
        try:
            text = content[:2000].decode("utf-8", errors="ignore")
            if "word/" in text:
                return ".docx"
            if "xl/" in text:
                return ".xlsx"
        except Exception:
            pass
        # 无法区分时回退为 .docx（最常见场景）
        return ".docx"
    return ".bin"


def normalize_payment_terms(data: dict) -> None:
    """归一化 payment_terms：installment_name→name, due_date→condition 兜底, 缺 name 补序号。"""
    if not isinstance(data, dict):
        return
    terms = data.get("payment_terms")
    if not isinstance(terms, list):
        return
    normalized = []
    for idx, t in enumerate(terms, 1):
        if not isinstance(t, dict):
            normalized.append(t)
            continue
        nt = dict(t)
        if "name" not in nt and "installment_name" in nt:
            nt["name"] = nt.pop("installment_name")
        if not nt.get("name"):
            nt["name"] = f"第 {idx} 期"
        if not nt.get("condition") and nt.get("due_date"):
            nt["condition"] = str(nt["due_date"])
        normalized.append(nt)
    data["payment_terms"] = normalized


# ─── VL / 文本模型调用（纯函数） ───

def make_text_extraction_prompt(base_prompt: str) -> str:
    """从 prompt 中移除 full_text 转录要求。
    文本 PDF 已由 PyMuPDF 提取全文，无需让 LLM 再转录一遍，节省 ~30-40s 生成时间。
    """
    lines = base_prompt.split("\n")
    result = []
    for line in lines:
        if "full_text" in line:
            continue
        result.append(line)
    return "\n".join(result)


def call_vl_model(file_bytes: bytes, mime: str, prompt: str) -> dict:
    """调用 DashScope VL 模型分析图片，返回结构化 JSON dict。"""
    image_base64 = base64.b64encode(file_bytes).decode()
    payload = {
        "model": settings.DASHSCOPE_VISION_MODEL,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{image_base64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        "temperature": 0.1,
        "max_tokens": 4096,
    }
    headers = {
        "Authorization": f"Bearer {settings.DASHSCOPE_API_KEY}",
        "Content-Type": "application/json",
    }
    with httpx.Client(timeout=120.0) as client:
        response = client.post(
            f"{settings.DASHSCOPE_BASE_URL}/chat/completions",
            json=payload, headers=headers,
        )
    if response.status_code != 200:
        raise RuntimeError(f"VL API 错误: {response.text}")

    content = response.json()["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"raw": content}


def call_text_model(text: str, prompt: str) -> dict:
    """调用 DeepSeek 文本模型分析文本，返回结构化 JSON dict。"""
    import time
    start = time.perf_counter()
    payload = {
        "model": settings.DEEPSEEK_AGENT_MODEL,
        "messages": [{"role": "user", "content": f"{prompt}\n\n以下是文件的文字内容，请提取结构化信息：\n\n{text[:8000]}"}],
        "temperature": 0.1,
        "max_tokens": 4096,
    }
    headers = {
        "Authorization": f"Bearer {settings.DEEPSEEK_API_KEY}",
        "Content-Type": "application/json",
    }
    with httpx.Client(timeout=120.0) as client:
        response = client.post(
            f"{settings.DEEPSEEK_BASE_URL}/chat/completions",
            json=payload, headers=headers,
        )
    if response.status_code != 200:
        raise RuntimeError(f"DashScope API 错误: {response.text}")

    content = response.json()["choices"][0]["message"]["content"]
    elapsed = time.perf_counter() - start
    logger.info("call_text_model: elapsed=%.2fs, text_len=%d, model=%s", elapsed, len(text), settings.DEEPSEEK_AGENT_MODEL)
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"raw": content}


def render_pdf_page_to_image(file_path: str, page_num: int = 0, dpi: int = 100) -> bytes:
    """将 PDF 指定页渲染为 PNG bytes"""
    import fitz
    doc = fitz.open(file_path)
    try:
        pix = doc[page_num].get_pixmap(dpi=dpi)
        return pix.tobytes("png")
    finally:
        doc.close()


def extract_pdf_text(file_path: str) -> str:
    """提取 PDF 所有页面文本"""
    import fitz
    doc = fitz.open(file_path)
    try:
        texts = []
        for page in doc:
            t = page.get_text().strip()
            if t:
                texts.append(t)
        return "\n\n".join(texts)
    finally:
        doc.close()


def extract_word_text(file_path: str) -> str:
    """提取 Word (.docx) 文档文本"""
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if paragraphs:
            return "\n".join(paragraphs)[:10000]
    except ImportError:
        pass
    return ""


def extract_excel_text(file_path: str) -> str:
    """提取 Excel (.xlsx) 表格数据"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows.append(f"[工作表: {sheet_name}]")
            count = 0
            for row in ws.iter_rows():
                vals = [str(c.value) if c.value is not None else "" for c in row]
                line = "\t".join(vals)
                if line.strip():
                    rows.append(line)
                    count += 1
                if count >= 200:
                    rows.append("... (仅前 200 行)")
                    break
        wb.close()
        return "\n".join(rows)[:10000]
    except ImportError:
        return ""
    except Exception:
        return ""


def is_docx(content: bytes) -> bool:
    """通过文件头检测 .docx（ZIP 签名 + 内部 [Content_Types].xml）"""
    if content[:4] != b"PK\x03\x04":
        return False
    return b"[Content_Types]" in content[:2000]


def is_xlsx(content: bytes) -> bool:
    """通过文件头检测 .xlsx（ZIP 签名 + 内部 xl/ 目录）"""
    if content[:4] != b"PK\x03\x04":
        return False
    return b"xl/" in content[:2000]


def extract_plain_text(file_bytes: bytes) -> str:
    """尝试多种编码提取纯文本内容"""
    for encoding in ("utf-8", "gbk", "gb2312", "utf-16"):
        try:
            return file_bytes[:20000].decode(encoding)[:10000]
        except (UnicodeDecodeError, UnicodeError):
            continue
    return ""
