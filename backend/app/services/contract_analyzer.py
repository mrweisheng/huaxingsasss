"""
合同文件分析器
从 tools.py 提取的共享 VL 分析逻辑，供 Agent 工具和页面端点复用。
职责：文件类型检测 → 图片压缩 → VL/文本模型调用 → 结构化输出 + 重复检测 + 自动创建付款。
不涉及 Redis 缓存、session 上下文。
"""
import base64
import io
import json
import os
import logging
import shutil
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import re

import httpx
from sqlalchemy.orm import Session

from app.config import settings
from app.models.contract import Contract
from app.models.user import User
from app.ai.prompts import CONTRACT_ANALYSIS_PROMPT
from app.utils.file_utils import calculate_file_hash
from app.services.contract_service import ContractService
from app.services.payment_service import PaymentService

logger = logging.getLogger(__name__)

# 图片压缩参数（与 tools.py 一致）
MAX_IMAGE_DIMENSION = 1600
JPEG_QUALITY = 85


def _compress_image(file_bytes: bytes, mime: str) -> tuple:
    """压缩图片：缩放到 MAX_IMAGE_DIMENSION 内 + JPEG 质量 85。"""
    from PIL import Image

    try:
        img = Image.open(io.BytesIO(file_bytes))
    except Exception:
        return file_bytes, mime

    w, h = img.size
    if max(w, h) <= MAX_IMAGE_DIMENSION and mime == "image/jpeg" and len(file_bytes) < 500_000:
        return file_bytes, mime

    if max(w, h) > MAX_IMAGE_DIMENSION:
        ratio = MAX_IMAGE_DIMENSION / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

    buf = io.BytesIO()
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    img.save(buf, format="JPEG", quality=JPEG_QUALITY)

    compressed = buf.getvalue()
    logger.info("图片压缩: %dx%d %s %.0fKB → %.0fKB", w, h, mime, len(file_bytes)/1024, len(compressed)/1024)
    return compressed, "image/jpeg"


def _detect_image_mime(header: bytes) -> str:
    """读取文件头判断图片 MIME 类型"""
    if header[:4] == b"\x89PNG":
        return "image/png"
    if header[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if header[:4] == b"GIF8":
        return "image/gif"
    if header[:4] == b"RIFF" and len(header) > 11 and header[8:12] == b"WEBP":
        return "image/webp"
    if header[:2] == b"BM":
        return "image/bmp"
    return ""


def _guess_extension(content: bytes) -> str:
    """通过文件头判断扩展名"""
    if content[:4] == b"%PDF":
        return ".pdf"
    if content[:3] == b"\xff\xd8\xff":
        return ".jpg"
    if content[:4] == b"\x89PNG":
        return ".png"
    if content[:4] == b"GIF8":
        return ".gif"
    if content[:4] == b"RIFF" and len(content) > 11 and content[8:12] == b"WEBP":
        return ".webp"
    return ".bin"


# ─── VL / 文本模型调用（纯函数） ───

def _make_text_extraction_prompt(base_prompt: str) -> str:
    """从 prompt 中移除 full_text 转录要求。
    文本 PDF 已由 PyMuPDF 提取全文，无需让 LLM 再转录一遍，节省 ~30-40s 生成时间。
    """
    lines = base_prompt.split("\n")
    result = []
    for line in lines:
        if "full_text" in line:
            continue
        result.append(line)
    return "\n".join(result)


def _call_vl_model(file_bytes: bytes, mime: str, prompt: str) -> dict:
    """调用 DashScope VL 模型分析图片，返回结构化 JSON dict。"""
    image_base64 = base64.b64encode(file_bytes).decode()
    payload = {
        "model": settings.DASHSCOPE_VISION_MODEL,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{image_base64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        "temperature": 0.1,
        "max_tokens": 4096,
    }
    headers = {
        "Authorization": f"Bearer {settings.DASHSCOPE_API_KEY}",
        "Content-Type": "application/json",
    }
    with httpx.Client(timeout=120.0) as client:
        response = client.post(
            f"{settings.DASHSCOPE_BASE_URL}/chat/completions",
            json=payload, headers=headers,
        )
    if response.status_code != 200:
        raise RuntimeError(f"VL API 错误: {response.text}")

    content = response.json()["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"raw": content}


def _call_text_model(text: str, prompt: str) -> dict:
    """调用 DashScope 文本模型分析文本，返回结构化 JSON dict。"""
    payload = {
        "model": settings.DASHSCOPE_TEXT_MODEL,
        "messages": [{"role": "user", "content": f"{prompt}\n\n以下是合同文件的文字内容，请提取结构化信息：\n\n{text[:8000]}"}],
        "temperature": 0.1,
        "max_tokens": 4096,
    }
    headers = {
        "Authorization": f"Bearer {settings.DASHSCOPE_API_KEY}",
        "Content-Type": "application/json",
    }
    with httpx.Client(timeout=120.0) as client:
        response = client.post(
            f"{settings.DASHSCOPE_BASE_URL}/chat/completions",
            json=payload, headers=headers,
        )
    if response.status_code != 200:
        raise RuntimeError(f"DashScope API 错误: {response.text}")

    content = response.json()["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"raw": content}


def _render_pdf_page_to_image(file_path: str, page_num: int = 0, dpi: int = 150) -> bytes:
    """将 PDF 指定页渲染为 PNG bytes"""
    import fitz
    doc = fitz.open(file_path)
    try:
        pix = doc[page_num].get_pixmap(dpi=dpi)
        return pix.tobytes("png")
    finally:
        doc.close()


def _extract_pdf_text(file_path: str) -> str:
    """提取 PDF 所有页面文本"""
    import fitz
    doc = fitz.open(file_path)
    try:
        texts = []
        for page in doc:
            t = page.get_text().strip()
            if t:
                texts.append(t)
        return "\n\n".join(texts)
    finally:
        doc.close()


def _extract_word_text(file_path: str) -> str:
    """提取 Word (.docx) 文档文本"""
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if paragraphs:
            return "\n".join(paragraphs)[:10000]
    except ImportError:
        pass
    return ""


def _extract_excel_text(file_path: str) -> str:
    """提取 Excel (.xlsx) 表格数据"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows.append(f"[工作表: {sheet_name}]")
            count = 0
            for row in ws.iter_rows():
                vals = [str(c.value) if c.value is not None else "" for c in row]
                line = "\t".join(vals)
                if line.strip():
                    rows.append(line)
                    count += 1
                if count >= 200:
                    rows.append("... (仅前 200 行)")
                    break
        wb.close()
        return "\n".join(rows)[:10000]
    except ImportError:
        return ""
    except Exception:
        return ""


def _is_docx(content: bytes) -> bool:
    """通过文件头检测 .docx（ZIP 签名 + 内部 [Content_Types].xml）"""
    if content[:4] != b"PK\x03\x04":
        return False
    return b"[Content_Types]" in content[:2000]


def _is_xlsx(content: bytes) -> bool:
    """通过文件头检测 .xlsx（ZIP 签名 + 内部 xl/ 目录）"""
    if content[:4] != b"PK\x03\x04":
        return False
    return b"xl/" in content[:2000]


def _extract_plain_text(file_bytes: bytes) -> str:
    """尝试多种编码提取纯文本内容"""
    for encoding in ("utf-8", "gbk", "gb2312", "utf-16"):
        try:
            return file_bytes[:20000].decode(encoding)[:10000]
        except (UnicodeDecodeError, UnicodeError):
            continue
    return ""


# ─── 归一化 payment_terms ───

def normalize_payment_terms(data: dict) -> None:
    """归一化 payment_terms：installment_name→name, due_date→condition 兜底, 缺 name 补序号。"""
    if not isinstance(data, dict):
        return
    terms = data.get("payment_terms")
    if not isinstance(terms, list):
        return
    normalized = []
    for idx, t in enumerate(terms, 1):
        if not isinstance(t, dict):
            normalized.append(t)
            continue
        nt = dict(t)
        if "name" not in nt and "installment_name" in nt:
            nt["name"] = nt.pop("installment_name")
        if not nt.get("name"):
            nt["name"] = f"第 {idx} 期"
        if not nt.get("condition") and nt.get("due_date"):
            nt["condition"] = str(nt["due_date"])
        normalized.append(nt)
    data["payment_terms"] = normalized


# ─── 自动创建付款记录 ───

def auto_create_payments_from_terms(
    contract: Contract,
    contract_data: dict,
    db: Session,
    user_id: int,
) -> list:
    """根据 payment_terms 中标记为已支付的条款，自动创建付款记录。"""
    logger.info("自动付款创建开始: contract_id=%d", contract.id)
    if not isinstance(contract_data, dict):
        return []

    payment_terms = contract_data.get("payment_terms", [])
    if not payment_terms:
        return []

    paid_keywords = ["已付", "已缴纳", "付清", "已到账", "已收", "已支付"]
    auto_payments = []

    for idx, term in enumerate(payment_terms, 1):
        is_paid_field = term.get("is_paid")
        condition = (term.get("condition") or "").lower()
        is_paid_by_keywords = any(kw in condition for kw in paid_keywords)
        is_paid_term = (is_paid_field is True) or (is_paid_field is None and is_paid_by_keywords)

        if not is_paid_term:
            continue

        try:
            term_amount = float(term.get("amount", 0))
        except (TypeError, ValueError):
            continue
        if term_amount <= 0:
            continue

        try:
            installment_number = PaymentService.get_next_installment_number(db, contract.id, "income")
            payment = PaymentService.create_payment_with_exchange_rate(
                db=db,
                contract_id=contract.id,
                installment_number=installment_number,
                currency=contract.currency,
                amount=Decimal(str(term_amount)),
                paid_date=contract.signed_date or date.today(),
                payment_method="unknown",
                receipt_image_path=None,
                notes="合同标注已付，待补充凭证",
                created_by=user_id,
                type="income",
                installment_name=term.get("name"),
            )
            auto_payments.append({
                "payment_id": payment.id,
                "installment_number": idx,
                "installment_name": term.get("name"),
                "amount": term_amount,
                "currency": contract.currency,
                "status": payment.status,
            })
        except Exception as e:
            logger.warning("自动创建付款失败: term=%s, error=%s", term.get("name"), e)
            auto_payments.append({
                "error": str(e),
                "installment_name": term.get("name"),
                "amount": term.get("amount"),
            })

    return auto_payments


# ─── 核心分析入口 ───

class ContractAnalyzer:
    """合同文件分析器——纯分析逻辑，不涉及缓存/session"""

    @staticmethod
    def analyze_file(
        file_path: str,
        db: Session,
        user_id: Optional[int] = None,
        skip_duplicate_check: bool = False,
    ) -> dict:
        """
        分析合同文件（图片/PDF/Word/Excel），返回结构化分析结果。

        Args:
            file_path: 文件绝对路径
            db: 数据库 session（用于重复检测）
            user_id: 用户 ID（仅用于日志）
            skip_duplicate_check: 跳过重复检测（用户确认"仍然创建"时传 True）

        Returns:
            dict: {
                "success": bool,
                "data": {...},           # 结构化分析结果
                "file_type": str,        # "image" | "pdf" | "word" | "excel"
                "file_hash": str,        # 文件哈希
                "duplicate_detected": bool,
                "existing_contract": {...} | None,
            }
        """
        # 读取文件
        with open(file_path, "rb") as f:
            header = f.read(12)
            f.seek(0)
            file_bytes = f.read()

        file_hash = calculate_file_hash(file_bytes)

        # 重复检测（可跳过）
        if not skip_duplicate_check:
            existing_contract = db.query(Contract).filter(
                Contract.file_hash == file_hash,
                Contract.is_deleted == False,
            ).first()
            if existing_contract:
                return {
                    "success": True,
                    "duplicate_detected": True,
                    "data": None,
                    "file_hash": file_hash,
                    "file_type": None,
                    "message": "该文件已在系统中存在对应的合同记录",
                    "existing_contract": {
                        "id": existing_contract.id,
                        "contract_number": existing_contract.contract_number,
                        "title": existing_contract.title,
                        "status": existing_contract.status,
                        "total_amount": float(existing_contract.total_amount) if existing_contract.total_amount else 0,
                        "currency": existing_contract.currency,
                        "customer_name": existing_contract.customer.name if existing_contract.customer else None,
                    },
                }

        # 判断文件类型并分析
        mime = _detect_image_mime(header)

        if mime:
            # 图片 → VL 模型
            compressed_bytes, compressed_mime = _compress_image(file_bytes, mime)
            structured = _call_vl_model(compressed_bytes, compressed_mime, CONTRACT_ANALYSIS_PROMPT)
            file_type = "image"
        elif header[:4] == b"%PDF":
            # PDF → 多策略
            full_text = _extract_pdf_text(file_path)
            if full_text.strip():
                logger.info("PDF 有文本，使用文本模型解析: text_len=%d", len(full_text))
                prompt = _make_text_extraction_prompt(CONTRACT_ANALYSIS_PROMPT)
                structured = _call_text_model(full_text, prompt)
                if isinstance(structured, dict):
                    structured["full_text"] = full_text
            else:
                logger.info("PDF 无文本（扫描件），渲染为图片走 VL")
                img_bytes = _render_pdf_page_to_image(file_path)
                img_bytes, _ = _compress_image(img_bytes, "image/png")
                structured = _call_vl_model(img_bytes, "image/png", CONTRACT_ANALYSIS_PROMPT)
            file_type = "pdf"
        else:
            # 尝试 Word / Excel
            text_content = ""
            # 根据文件扩展名判断（agent/upload 不保留扩展名，所以也尝试内容检测）
            file_name = os.path.basename(file_path)
            if file_name.endswith(".docx") or _is_docx(file_bytes):
                text_content = _extract_word_text(file_path)
                file_type = "word"
            elif file_name.endswith(".xlsx") or _is_xlsx(file_bytes):
                text_content = _extract_excel_text(file_path)
                file_type = "excel"
            else:
                # 最后尝试纯文本编码检测
                text_content = _extract_plain_text(file_bytes)
                file_type = "text"

            if not text_content.strip():
                return {
                    "success": False,
                    "error": "无法提取文件内容，仅支持图片（JPEG/PNG）、PDF、Word（.docx）、Excel（.xlsx）格式",
                    "file_hash": file_hash,
                    "file_type": None,
                    "data": None,
                    "duplicate_detected": False,
                    "existing_contract": None,
                }

            # 文本内容 → 百炼 DashScope 文本模型（qwen-plus）
            structured = _call_text_model(text_content, CONTRACT_ANALYSIS_PROMPT)

        # 归一化 payment_terms
        normalize_payment_terms(structured)

        return {
            "success": True,
            "duplicate_detected": False,
            "data": structured,
            "file_hash": file_hash,
            "file_type": file_type,
            "existing_contract": None,
            "message": None,
        }

    @staticmethod
    def resolve_file_path(file_id: str, user_id: int) -> Optional[str]:
        """解析文件路径：优先用户隔离路径，回退全局路径。"""
        candidates = [
            os.path.join(settings.TEMP_UPLOAD_DIR, str(user_id), file_id),
            os.path.join(settings.TEMP_UPLOAD_DIR, file_id),
        ]
        return next((p for p in candidates if os.path.exists(p)), None)

    @staticmethod
    def copy_to_contract_dir(temp_file_path: str, contract_number: str) -> str:
        """
        将临时文件复制到合同永久目录。
        返回相对路径（如 2026/06/HT20260603xxxx.jpg）。
        """
        with open(temp_file_path, "rb") as f:
            content = f.read()
        ext = _guess_extension(content)
        year_month = datetime.now().strftime("%Y/%m")
        target_dir = Path(settings.CONTRACT_UPLOAD_DIR) / year_month
        target_dir.mkdir(parents=True, exist_ok=True)
        target_filename = f"{contract_number}{ext}"
        target_path = target_dir / target_filename
        shutil.copy2(temp_file_path, str(target_path))
        return str(Path(year_month) / target_filename)
